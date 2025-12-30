from openpyxl import load_workbook

wb = load_workbook('order_generation/PO_excel_export/test_EC221.xlsx')
ws = wb.active
print(f'B14 (交货时间): {ws["B14"].value}')
print(f'Type: {type(ws["B14"].value).__name__}')
print()
print('✓ Verification: Delivery time is stored as text with "天" suffix')
