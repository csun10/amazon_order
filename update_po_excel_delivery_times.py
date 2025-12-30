"""
One-time script to update 交货时间 (delivery time) in PO_excel files
Converts date values to "XX天" format based on supplier information
"""

import os
import glob
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText

def get_supplier_delivery_days(supplier_name):
    """Determine delivery days based on supplier name"""
    if '印刷厂' in supplier_name:
        return 7
    elif supplier_name in [
        '宁波泰丰机械有限公司',
        '阳江骏业工贸有限公司', 
        '宁波瑾秀制刷科技有限公司',
        '宁波市海曙硕丰塑料五金制品有限公司'
    ]:
        return 45
    else:
        return 15

def get_cell_value(ws, row, col):
    """Get cell value as string, handling formulas and rich text"""
    cell = ws.cell(row=row, column=col)
    
    if cell.value is None:
        return ""
    elif isinstance(cell.value, str):
        return cell.value.strip()
    elif isinstance(cell.value, CellRichText):
        return str(cell.value).strip()
    else:
        return str(cell.value).strip()

def update_po_excel_delivery_time(excel_path):
    """Update delivery time in a PO Excel file"""
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # Get supplier name from B3
        supplier_value = get_cell_value(ws, 3, 2)  # B3
        
        # Get current delivery time from B14
        current_delivery = get_cell_value(ws, 14, 2)  # B14
        
        # Skip if already has '天' suffix
        if current_delivery and current_delivery.endswith('天'):
            return False, current_delivery, "Already has 天 suffix"
        
        # Determine new delivery days based on supplier
        new_days = get_supplier_delivery_days(supplier_value)
        new_value = f"{new_days}天"
        
        # Update cell B14
        ws.cell(row=14, column=2).value = new_value
        
        # Save the workbook
        wb.save(excel_path)
        
        return True, new_value, f"Updated from '{current_delivery}' (Supplier: {supplier_value[:30]})"
        
    except Exception as e:
        return False, None, f"Error: {e}"

def main():
    # Get all Excel files in PO_excel folder
    po_excel_dir = Path('order_generation/PO_excel')
    
    if not po_excel_dir.exists():
        print(f"Error: {po_excel_dir} does not exist")
        return
    
    excel_files = list(po_excel_dir.glob('*.xlsx'))
    
    # Filter out temporary files
    excel_files = [f for f in excel_files if not f.name.startswith('~$')]
    
    print("=" * 80)
    print("One-Time Update: PO_excel 交货时间 (Delivery Time)")
    print("=" * 80)
    print(f"Found {len(excel_files)} Excel files in {po_excel_dir}")
    print()
    
    updated_count = 0
    skipped_count = 0
    error_count = 0
    
    for excel_file in sorted(excel_files):
        success, new_value, message = update_po_excel_delivery_time(excel_file)
        
        if success:
            updated_count += 1
            print(f"✓ {excel_file.name:40} → {new_value:6} ({message})")
        else:
            if "Already has" in message:
                skipped_count += 1
                print(f"○ {excel_file.name:40} → {new_value:6} (Skipped)")
            else:
                error_count += 1
                print(f"✗ {excel_file.name:40} {message}")
    
    print()
    print("=" * 80)
    print("Summary:")
    print(f"  Updated:  {updated_count} files")
    print(f"  Skipped:  {skipped_count} files (already correct)")
    print(f"  Errors:   {error_count} files")
    print("=" * 80)
    print()

if __name__ == "__main__":
    main()
