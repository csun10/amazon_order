#!/usr/bin/env python3
"""Batch verification script to process all SKUs with sales > 10 in 30 days."""

import openpyxl
import sys
from pathlib import Path
from direct_sku_to_json import generate_factory_jsons
from fill_po_import import fill_po_import_for_order
import json

def read_listing_file(excel_path):
    """Read listing Excel and extract SKUs with sales > 10 in 30 days."""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    print(f"Total rows in listing: {ws.max_row}")
    
    # Extract data (assuming column 10 is SKU, column 13 is 30-day sales)
    skus_to_process = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        try:
            sku = row[9].value  # Column J (index 9)
            sales_30_days = row[12].value  # Column M (index 12)
            
            if sku and sales_30_days:
                # Convert to number if needed
                if isinstance(sales_30_days, str):
                    sales_30_days = float(sales_30_days.replace(',', ''))
                
                if isinstance(sales_30_days, (int, float)) and sales_30_days > 10:
                    skus_to_process.append({
                        'sku': str(sku).strip(),
                        'sales': int(sales_30_days),
                        'row': row_idx
                    })
        except Exception as e:
            print(f"Warning: Error processing row {row_idx}: {e}")
            continue
    
    return skus_to_process


def run_verification(listing_file):
    """Run order generation for all qualifying SKUs."""
    listing_path = Path(listing_file)
    
    if not listing_path.exists():
        print(f"Error: Listing file not found: {listing_file}")
        return 1
    
    print(f"Reading listing file: {listing_file}")
    skus_to_process = read_listing_file(listing_path)
    
    print(f"\n{'='*80}")
    print(f"Found {len(skus_to_process)} SKUs with 30-day sales > 10")
    print(f"{'='*80}\n")
    
    # Display all SKUs to be processed
    for idx, sku_info in enumerate(skus_to_process, 1):
        print(f"{idx:3d}. SKU: {sku_info['sku']:30s} Sales: {sku_info['sales']:5d}")
    
    print(f"\n{'='*80}")
    print(f"Starting batch order generation...")
    print(f"{'='*80}\n")
    
    # Process each SKU one by one
    results = []
    template_dir = Path(__file__).parent / "json_template"
    
    for idx, sku_info in enumerate(skus_to_process, 1):
        sku = sku_info['sku']
        sales = sku_info['sales']
        
        print(f"\n{'='*80}")
        print(f"Processing {idx}/{len(skus_to_process)}: {sku} (Sales: {sales})")
        print(f"{'='*80}")
        
        # Check if template exists
        template_path = template_dir / f"{sku}.json"
        if not template_path.exists():
            print(f"[SKIPPED] Template not found for {sku}")
            results.append({
                'sku': sku,
                'sales': sales,
                'status': 'SKIPPED',
                'reason': 'Template not found'
            })
            continue
        
        try:
            # Generate order with quantity = sales
            # Replace underscores with dashes for ERP compatibility
            order_name = f"verify-{sku}".replace("_", "-")
            pairs = {sku: sales}
            
            print(f"Generating order: {order_name}")
            print(f"Quantity: {sales}")
            
            json_paths = generate_factory_jsons(pairs, order_name)
            
            # Generate PO import
            try:
                po_import_path = fill_po_import_for_order(order_name, warehouse="默认仓库")
                print(f"[SUCCESS] Generated PO import: {po_import_path}")
                
                results.append({
                    'sku': sku,
                    'sales': sales,
                    'status': 'SUCCESS',
                    'json_files': [str(p) for p in json_paths],
                    'po_import': str(po_import_path)
                })
            except Exception as e:
                print(f"[WARNING] PO import generation failed: {e}")
                results.append({
                    'sku': sku,
                    'sales': sales,
                    'status': 'PARTIAL',
                    'reason': f'PO import failed: {e}',
                    'json_files': [str(p) for p in json_paths]
                })
            
        except Exception as e:
            print(f"[ERROR] processing {sku}: {e}")
            results.append({
                'sku': sku,
                'sales': sales,
                'status': 'ERROR',
                'reason': str(e)
            })
    
    # Print summary
    print(f"\n{'='*80}")
    print(f"VERIFICATION SUMMARY")
    print(f"{'='*80}\n")
    
    success_count = sum(1 for r in results if r['status'] == 'SUCCESS')
    partial_count = sum(1 for r in results if r['status'] == 'PARTIAL')
    skipped_count = sum(1 for r in results if r['status'] == 'SKIPPED')
    error_count = sum(1 for r in results if r['status'] == 'ERROR')
    
    print(f"Total SKUs processed: {len(results)}")
    print(f"[OK] Success: {success_count}")
    print(f"[WARN] Partial: {partial_count}")
    print(f"[SKIP] Skipped: {skipped_count}")
    print(f"[ERROR] Errors: {error_count}")
    
    # Save results to JSON
    results_path = Path(__file__).parent / "verification_output" / "batch_verification_results.json"
    results_path.parent.mkdir(exist_ok=True)
    
    with open(results_path, 'w', encoding='utf-8') as f:
        json.dump({
            'summary': {
                'total': len(results),
                'success': success_count,
                'partial': partial_count,
                'skipped': skipped_count,
                'error': error_count
            },
            'details': results
        }, f, ensure_ascii=False, indent=2)
    
    print(f"\n[OK] Results saved to: {results_path}")
    
    # List failed SKUs
    if skipped_count > 0 or error_count > 0:
        print(f"\n{'='*80}")
        print(f"FAILED SKUs:")
        print(f"{'='*80}\n")
        
        for r in results:
            if r['status'] in ['SKIPPED', 'ERROR']:
                print(f"  {r['sku']:30s} - {r['status']}: {r.get('reason', 'Unknown')}")
    
    return 0


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python batch_verification.py <listing_excel_file>")
        sys.exit(1)
    
    listing_file = sys.argv[1]
    sys.exit(run_verification(listing_file))
