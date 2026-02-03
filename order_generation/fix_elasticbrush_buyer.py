# -*- coding: utf-8 -*-
import sys, io
from openpyxl import load_workbook
from pathlib import Path
import json

if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

ROOT = Path(__file__).resolve().parent
BUYER_JIXIU = "宁波集秀美容科技有限公司"

# Fix Elasticbrush01
sku = "Elasticbrush01"
excel_path = ROOT / "PO_excel_template" / f"{sku}.xlsx"
json_path = ROOT / "json_template" / f"{sku}.json"

print(f"Fixing {sku}...")

# Update Excel B69
wb = load_workbook(excel_path)
ws = wb.active

print(f"  Current B69: {ws['B69'].value}")
ws['B69'] = BUYER_JIXIU
wb.save(excel_path)
print(f"  Updated B69: {BUYER_JIXIU}")

# Update JSON
from excel_to_json_template import ExcelToJsonConverter
conv = ExcelToJsonConverter()

json_data = conv.convert_excel_to_json(excel_path)

with open(json_path, 'w', encoding='utf-8') as f:
    json.dump(json_data, f, ensure_ascii=False, indent=2)

# Verify
with open(json_path, 'r', encoding='utf-8') as f:
    data = json.load(f)

print(f"  JSON B69 (footer.buyer): {data.get('footer', {}).get('buyer')}")
print(f"\n✓ Fixed {sku}!")
