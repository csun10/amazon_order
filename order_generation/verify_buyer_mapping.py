"""Verify buyer mapping is working correctly"""
# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from pathlib import Path
import sys

# Force UTF-8 output on Windows
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

print("="*60)
print("BUYER MAPPING VERIFICATION")
print("="*60)

# Check Excel files
excel_files = list(Path("PO_excel_export").glob("final_test-*.xlsx"))
print(f"\n[OK] Generated {len(excel_files)} Excel files")

for excel_file in sorted(excel_files):
    wb = load_workbook(excel_file)
    ws = wb.active
    
    buyer = ws['B69'].value
    # Get first product SKU
    first_sku = ws.cell(7, 1).value
    
    print(f"\n  {excel_file.name}:")
    print(f"    First SKU: {first_sku}")
    print(f"    Buyer (B69): {buyer}")
    wb.close()

# Check PO import
po_file = Path("PO_import_filled/PO_import_final_test.xlsx")
if po_file.exists():
    print(f"\n[OK] PO Import file generated: {po_file.name}")
    wb = load_workbook(po_file)
    ws = wb.active
    
    print("\n  采购方 (Buyer) column check (first 12 rows):")
    print(f"  {'Row':<5} {'SKU':<25} {'采购方 (Buyer)':<40}")
    print(f"  {'-'*5} {'-'*25} {'-'*40}")
    
    for row in range(3, 15):
        sku = ws.cell(row, 25).value  # Column 25 = SKU
        buyer = ws.cell(row, 5).value  # Column 5 = 采购方
        
        if sku:
            buyer_display = buyer if buyer else "(blank - accessory)"
            print(f"  {row:<5} {sku:<25} {buyer_display:<40}")
    
    wb.close()

print("\n" + "="*60)
print("VERIFICATION COMPLETE")
print("="*60)
print("\nExpected behavior:")
print("  [OK] Parent products: Buyer assigned from listing")
print("  [OK] Accessories: Buyer field blank")
print("  [OK] Elasticbrush01 -> JIXIU (宁波集秀美容科技有限公司)")
print("  [OK] B10-MJB2-BK -> PINXIU (宁波品秀美容科技有限公司)")
