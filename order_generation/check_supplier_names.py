#!/usr/bin/env python3
"""Check PO excel templates for supplier names that don't match the reference supplier list.

This script reads the reference supplier list and checks all PO excel templates
to find any with supplier names that don't exactly match (including blank ones).
"""

import openpyxl
from pathlib import Path
from typing import Set, List, Dict
import json

# Paths
ROOT = Path(__file__).resolve().parent
REFERENCE_FILE = ROOT / "docs" / "Supplier20260202170516-876874265653952512.xlsx"
PO_EXCEL_TEMPLATE_DIR = ROOT / "PO_excel_template"
JSON_TEMPLATE_DIR = ROOT / "json_template"

def read_supplier_reference_list(xlsx_path: Path) -> Set[str]:
    """Read the supplier reference list from Excel file.
    
    Returns:
        Set of exact supplier names from the reference file
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    
    suppliers = set()
    
    # Read suppliers from the file - check the structure first
    print(f"\n=== Reading reference file: {xlsx_path.name} ===")
    print(f"Sheet name: {ws.title}")
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
    
    # Check first few rows to understand structure
    print("\nFirst few rows (checking columns 1-5):")
    for row in range(1, min(6, ws.max_row + 1)):
        row_data = []
        for col in range(1, min(6, ws.max_column + 1)):
            val = ws.cell(row, col).value
            row_data.append(str(val) if val else "")
        print(f"Row {row}: {row_data}")
    
    # Based on the structure, supplier name should be in column 3 (column C)
    # Row 1 is header, data starts from row 2
    print("\nExtracting suppliers from column 3...")
    start_row = 2  # Start from row 2 (skip header)
    
    for row in range(start_row, ws.max_row + 1):
        supplier = ws.cell(row, 3).value  # Column 3 = supplier name
        if supplier and str(supplier).strip():
            supplier_str = str(supplier).strip()
            suppliers.add(supplier_str)
    
    wb.close()
    
    print(f"\n=== Found {len(suppliers)} unique suppliers in reference file ===")
    for supplier in sorted(suppliers):
        print(f"  - {supplier}")
    
    return suppliers

def check_excel_supplier(xlsx_path: Path) -> tuple:
    """Check supplier name in a PO excel template.
    
    Returns:
        Tuple of (supplier_name, cell_location)
    """
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        
        # Supplier is in cell B3 according to the JSON structure
        supplier_b3 = ws['B3'].value
        supplier = str(supplier_b3).strip() if supplier_b3 else ""
        
        wb.close()
        return (supplier, "B3")
    except Exception as e:
        return (f"ERROR: {e}", "")

def check_json_supplier(json_path: Path) -> str:
    """Check supplier name in a JSON template.
    
    Returns:
        Supplier name from JSON B3 value
    """
    try:
        with open(json_path, 'r', encoding='utf-8-sig') as f:
            data = json.load(f)
        
        supplier = data.get("cells", {}).get("B3", {}).get("value", "")
        return str(supplier).strip() if supplier else ""
    except Exception as e:
        return f"ERROR: {e}"

def main():
    """Main function."""
    print("=" * 80)
    print("SUPPLIER NAME VALIDATION CHECK")
    print("=" * 80)
    
    # Check if reference file exists
    if not REFERENCE_FILE.exists():
        print(f"\nERROR: Reference file not found: {REFERENCE_FILE}")
        return 1
    
    # Read reference supplier list
    valid_suppliers = read_supplier_reference_list(REFERENCE_FILE)
    
    if not valid_suppliers:
        print("\nWARNING: No suppliers found in reference file!")
        return 1
    
    # Check PO excel templates
    print(f"\n\n{'=' * 80}")
    print("CHECKING PO EXCEL TEMPLATES")
    print("=" * 80)
    
    if not PO_EXCEL_TEMPLATE_DIR.exists():
        print(f"\nERROR: PO excel template directory not found: {PO_EXCEL_TEMPLATE_DIR}")
        return 1
    
    excel_files = list(PO_EXCEL_TEMPLATE_DIR.glob("*.xlsx"))
    excel_files = [f for f in excel_files if not f.name.startswith('~$')]  # Skip temp files
    excel_files.sort()
    
    print(f"\nFound {len(excel_files)} Excel template files\n")
    
    mismatches = []
    blank_suppliers = []
    matches = []
    errors = []
    
    for excel_file in excel_files:
        supplier, cell = check_excel_supplier(excel_file)
        
        if supplier.startswith("ERROR:"):
            errors.append({
                'file': excel_file.name,
                'error': supplier
            })
        elif not supplier:
            blank_suppliers.append({
                'file': excel_file.name,
                'supplier': '(BLANK)',
                'cell': cell
            })
        elif supplier not in valid_suppliers:
            mismatches.append({
                'file': excel_file.name,
                'supplier': supplier,
                'cell': cell
            })
        else:
            matches.append({
                'file': excel_file.name,
                'supplier': supplier
            })
    
    # Check JSON templates
    print(f"\n\n{'=' * 80}")
    print("CHECKING JSON TEMPLATES")
    print("=" * 80)
    
    if not JSON_TEMPLATE_DIR.exists():
        print(f"\nERROR: JSON template directory not found: {JSON_TEMPLATE_DIR}")
        return 1
    
    json_files = list(JSON_TEMPLATE_DIR.glob("*.json"))
    json_files.sort()
    
    print(f"\nFound {len(json_files)} JSON template files\n")
    
    json_mismatches = []
    json_blank_suppliers = []
    json_matches = []
    json_errors = []
    
    for json_file in json_files:
        supplier = check_json_supplier(json_file)
        
        if supplier.startswith("ERROR:"):
            json_errors.append({
                'file': json_file.name,
                'error': supplier
            })
        elif not supplier:
            json_blank_suppliers.append({
                'file': json_file.name,
                'supplier': '(BLANK)'
            })
        elif supplier not in valid_suppliers:
            json_mismatches.append({
                'file': json_file.name,
                'supplier': supplier
            })
        else:
            json_matches.append({
                'file': json_file.name,
                'supplier': supplier
            })
    
    # Report results
    print(f"\n\n{'=' * 80}")
    print("RESULTS SUMMARY")
    print("=" * 80)
    
    print(f"\n=== PO EXCEL TEMPLATES ===")
    print(f"Total files checked: {len(excel_files)}")
    print(f"  [OK] Matching suppliers: {len(matches)}")
    print(f"  [X] Non-matching suppliers: {len(mismatches)}")
    print(f"  [!] Blank suppliers: {len(blank_suppliers)}")
    print(f"  [!] Errors: {len(errors)}")
    
    print(f"\n=== JSON TEMPLATES ===")
    print(f"Total files checked: {len(json_files)}")
    print(f"  [OK] Matching suppliers: {len(json_matches)}")
    print(f"  [X] Non-matching suppliers: {len(json_mismatches)}")
    print(f"  [!] Blank suppliers: {len(json_blank_suppliers)}")
    print(f"  [!] Errors: {len(json_errors)}")
    
    # Detailed reports
    if mismatches:
        print(f"\n\n{'=' * 80}")
        print("PO EXCEL TEMPLATES WITH NON-MATCHING SUPPLIER NAMES")
        print("=" * 80)
        for item in mismatches:
            print(f"\n  File: {item['file']}")
            print(f"  Supplier in file ({item['cell']}): '{item['supplier']}'")
            print(f"  Status: NOT FOUND in reference list")
    
    if blank_suppliers:
        print(f"\n\n{'=' * 80}")
        print("PO EXCEL TEMPLATES WITH BLANK SUPPLIER NAMES")
        print("=" * 80)
        for item in blank_suppliers:
            print(f"  - {item['file']} (cell {item['cell']})")
    
    if errors:
        print(f"\n\n{'=' * 80}")
        print("PO EXCEL TEMPLATES WITH ERRORS")
        print("=" * 80)
        for item in errors:
            print(f"  - {item['file']}: {item['error']}")
    
    if json_mismatches:
        print(f"\n\n{'=' * 80}")
        print("JSON TEMPLATES WITH NON-MATCHING SUPPLIER NAMES")
        print("=" * 80)
        for item in json_mismatches:
            print(f"\n  File: {item['file']}")
            print(f"  Supplier in file (B3): '{item['supplier']}'")
            print(f"  Status: NOT FOUND in reference list")
    
    if json_blank_suppliers:
        print(f"\n\n{'=' * 80}")
        print("JSON TEMPLATES WITH BLANK SUPPLIER NAMES")
        print("=" * 80)
        for item in json_blank_suppliers:
            print(f"  - {item['file']}")
    
    if json_errors:
        print(f"\n\n{'=' * 80}")
        print("JSON TEMPLATES WITH ERRORS")
        print("=" * 80)
        for item in json_errors:
            print(f"  - {item['file']}: {item['error']}")
    
    # Save results to file
    results = {
        'reference_file': str(REFERENCE_FILE),
        'valid_suppliers': sorted(list(valid_suppliers)),
        'excel_templates': {
            'total': len(excel_files),
            'matches': len(matches),
            'mismatches': mismatches,
            'blank': blank_suppliers,
            'errors': errors
        },
        'json_templates': {
            'total': len(json_files),
            'matches': len(json_matches),
            'mismatches': json_mismatches,
            'blank': json_blank_suppliers,
            'errors': json_errors
        }
    }
    
    output_file = ROOT / "supplier_validation_report.json"
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(results, f, indent=2, ensure_ascii=False)
    
    print(f"\n\n{'=' * 80}")
    print(f"Report saved to: {output_file}")
    print("=" * 80)
    
    # Return non-zero if there are any issues
    if mismatches or blank_suppliers or json_mismatches or json_blank_suppliers:
        return 1
    
    return 0

if __name__ == "__main__":
    import sys
    sys.exit(main())
