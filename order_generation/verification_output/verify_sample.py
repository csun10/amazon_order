#!/usr/bin/env python3
"""Quick verification script to check a sample generated Excel file."""

import openpyxl
import sys
from pathlib import Path

def verify_excel_file(filepath):
    """Verify the contents of a generated Excel file."""
    print(f"Verifying: {filepath}")
    print("=" * 70)
    
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    # Check key cells
    print(f"\nKey Information:")
    print(f"  Supplier (B3): {ws['B3'].value}")
    print(f"  Order Number (G3): {ws['G3'].value}")
    print(f"  Date (G4): {ws['G4'].value}")
    print(f"  Buyer (B69): {ws['B69'].value}")
    
    # Count products
    print(f"\nProducts in order:")
    count = 0
    for row in range(7, 50):
        sku = ws.cell(row, 1).value
        if sku and str(sku).strip():
            qty = ws.cell(row, 5).value
            price = ws.cell(row, 6).value
            name = ws.cell(row, 2).value
            count += 1
            print(f"  {count}. SKU: {sku}, Name: {name}, Qty: {qty}, Price: {price}")
    
    print(f"\nTotal products: {count}")
    print("=" * 70)
    print("✓ File verification complete\n")
    
    return count > 0

if __name__ == "__main__":
    # Verify a few sample files
    base_path = Path(__file__).parent.parent / "PO_excel_export"
    
    sample_files = [
        "verify_Elasticbrush01-1.xlsx",
        "verify_EEHB-NBB-1.xlsx",
        "verify_ST1122-1-1.xlsx"
    ]
    
    print("Sample File Verification Report")
    print("=" * 70)
    print()
    
    all_valid = True
    for filename in sample_files:
        filepath = base_path / filename
        if filepath.exists():
            try:
                valid = verify_excel_file(filepath)
                all_valid = all_valid and valid
            except Exception as e:
                print(f"ERROR verifying {filename}: {e}")
                all_valid = False
        else:
            print(f"File not found: {filename}")
            all_valid = False
    
    if all_valid:
        print("\n✓ All sample files passed verification!")
        sys.exit(0)
    else:
        print("\n✗ Some files failed verification")
        sys.exit(1)
