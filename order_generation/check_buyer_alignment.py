#!/usr/bin/env python3
"""Check 采购方 alignment in PO import files."""

import openpyxl
from pathlib import Path
import json

def check_buyer_alignment(po_file_path):
    """Check if accessories have consistent buyer with their parent products."""
    
    wb = openpyxl.load_workbook(po_file_path)
    ws = wb.active
    
    # Find column indices
    headers = [ws.cell(2, col).value for col in range(1, ws.max_column + 1)]
    
    sku_col = headers.index('*SKU') + 1 if '*SKU' in headers else None
    buyer_col = headers.index('采购方') + 1 if '采购方' in headers else None
    
    if not sku_col or not buyer_col:
        print("Required columns not found")
        return
    
    print(f"Checking: {po_file_path.name}")
    print("=" * 80)
    
    # Collect all rows
    rows = []
    for row_idx in range(3, ws.max_row + 1):
        sku = ws.cell(row_idx, sku_col).value
        buyer = ws.cell(row_idx, buyer_col).value
        if sku:
            rows.append({
                'row': row_idx,
                'sku': sku,
                'buyer': buyer
            })
    
    # Group by parent product (using accessory mapping)
    print(f"\n{'Row':<5} {'SKU':<30} {'采购方':<30}")
    print("-" * 80)
    
    for row in rows:
        buyer_display = row['buyer'] if row['buyer'] else "[EMPTY]"
        print(f"{row['row']:<5} {row['sku']:<30} {buyer_display:<30}")
    
    print("\n" + "=" * 80)
    
    # Check for inconsistencies
    buyers_found = set(r['buyer'] for r in rows if r['buyer'])
    print(f"\nUnique buyers found: {buyers_found}")
    
    # Count empty buyers
    empty_count = sum(1 for r in rows if not r['buyer'])
    filled_count = sum(1 for r in rows if r['buyer'])
    
    print(f"Filled 采购方: {filled_count}")
    print(f"Empty 采购方: {empty_count}")
    
    if empty_count > 0 and filled_count > 0:
        print("\n⚠ ISSUE: Some products have 采购方 while others don't")
        print("This suggests accessories are not inheriting buyer from parent products")
    
    return rows

if __name__ == "__main__":
    po_file = Path(__file__).parent / "PO_import_filled" / "PO_import_verify_ST1122-1.xlsx"
    
    if not po_file.exists():
        print(f"File not found: {po_file}")
        exit(1)
    
    check_buyer_alignment(po_file)
