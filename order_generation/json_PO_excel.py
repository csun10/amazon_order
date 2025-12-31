import sys
import json
import re
from datetime import datetime, timedelta
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image
    from openpyxl.cell.rich_text import TextBlock, CellRichText
    from openpyxl.cell.text import InlineFont
    from openpyxl.styles import Alignment
except ModuleNotFoundError as exc:  # pragma: no cover - dependency missing in tests
    raise SystemExit("openpyxl and pillow are required to run this script") from exc

PRODUCT_START_ROW = 7
COLUMN_MAP = {
    '产品编号': 'A',
    '产品图片': 'B',
    '描述': 'C',
    '数量/个': 'D',
    '单价': 'E',
    '包装方式': 'G',
}


def create_rich_text_from_json(description_data):
    """
    Convert JSON rich text format to openpyxl CellRichText object.
    
    Args:
        description_data: Always expects rich text format:
                         {
                           "type": "rich_text",
                           "content": [
                             {
                               "text": "text content",
                               "bold": true/false,
                               "color": "RRGGBB" (hex color)
                             },
                             ...
                           ]
                         }
                         
                         Plain text strings are automatically converted to 
                         single-block rich text format during processing.
    
    Returns:
        CellRichText object for multi-block content, or string for single plain block
    """
    # Convert plain text to rich text format for uniform processing
    if isinstance(description_data, str):
        description_data = {
            "type": "rich_text",
            "content": [{
                "text": description_data,
                "bold": False,
                "color": "000000"
            }]
        }
    
    # Handle rich text format (now the only expected format)
    if isinstance(description_data, dict) and description_data.get('type') == 'rich_text':
        content = description_data.get('content', [])
        
        # If single block with default formatting, return as plain text
        if (len(content) == 1 and 
            not content[0].get('bold', False) and 
            content[0].get('color', '000000') == '000000'):
            return content[0].get('text', '')
        
        # Create rich text for multiple blocks or formatted content
        rich_text = CellRichText()
        
        for content_block in content:
            text = content_block.get('text', '')
            bold = content_block.get('bold', False)
            color = content_block.get('color', '000000')  # Default to black
            
            # Ensure color is in proper format (6 hex digits)
            if not color.startswith('00') and len(color) == 6:
                color = '00' + color  # Add alpha channel for openpyxl
            
            # Create font with specified formatting
            font = InlineFont(b=bold, color=color)
            text_block = TextBlock(font, text)
            rich_text.append(text_block)
        
        return rich_text
    
    # Fallback to string representation for unknown formats
    return str(description_data)


def fill_workbook(template: Path, data: dict, json_filename: str = ""):
    """Return workbook filled with ``data`` using ``template``."""
    wb = load_workbook(template)
    ws = wb.active

    for addr, info in data.get('cells', {}).items():
        value = info.get('value', '')
        key = info.get('key', '')
        
        # Handle special date fields
        if key == '日期':
            # Fill with today's date
            value = datetime.now().strftime('%Y年%m月%d日')
        elif key == '交货时间':
            # Keep the delivery time as-is from JSON (e.g., "45天")
            # The value should already have "天" suffix from the JSON template
            value = str(value)
        elif key == '交货日期':
            # For 交货日期, calculate the actual delivery date
            original_value = str(value)
            # Look for numbers in the value (could be "15天", "30", "45", etc.)
            numbers = re.findall(r'\d+', original_value)
            if numbers:
                days_to_add = int(numbers[0])
                delivery_date = datetime.now() + timedelta(days=days_to_add)
                value = delivery_date.strftime('%Y年%m月%d日')
            else:
                # If no number found, default to 30 days from today
                delivery_date = datetime.now() + timedelta(days=30)
                value = delivery_date.strftime('%Y年%m月%d日')
        elif key == '订单号' and json_filename:
            # Always use the filename stem as the order number (e.g., "25AM027-1.json" -> "25AM027-1")
            order_number = Path(json_filename).stem
            value = order_number
        # Handle color cards (色卡) - insert color images
        elif key.startswith('色卡') and value:
            color_img_path = template.parent.parent / 'images' / 'colors' / f'{value}.png'
            if not color_img_path.exists():
                color_img_path = template.parent.parent / 'images' / 'colors' / f'{value}.jpg'
            
            if color_img_path.exists():
                try:
                    from PIL import Image as PILImage
                    # Verify image can be opened by Pillow
                    with PILImage.open(color_img_path) as pil_img:
                        pil_img.verify()
                        orig_width, orig_height = pil_img.size
                    
                    img = Image(str(color_img_path))
                    # Set appropriate size for color card (smaller than product images)
                    target_height_px = 60
                    scale = target_height_px / orig_height
                    img.height = target_height_px
                    img.width = int(orig_width * scale)
                    ws.add_image(img, addr)
                    # Don't set text value for image cells
                    continue
                except Exception as e:
                    value = f"[色卡图片错误] {value}: {e}"
            else:
                value = f"[色卡图片未找到] {value}"
        # Handle logos - insert logo images  
        elif key.startswith('Logo') and value:
            logo_img_path = template.parent.parent / 'images' / 'logos' / f'{value}'
            # If no extension provided, try common formats
            if not logo_img_path.suffix:
                for ext in ['.png', '.jpg', '.jpeg']:
                    test_path = template.parent.parent / 'images' / 'logos' / f'{value}{ext}'
                    if test_path.exists():
                        logo_img_path = test_path
                        break
            
            if logo_img_path.exists():
                try:
                    from PIL import Image as PILImage
                    # Verify image can be opened by Pillow
                    with PILImage.open(logo_img_path) as pil_img:
                        pil_img.verify()
                        orig_width, orig_height = pil_img.size
                    
                    img = Image(str(logo_img_path))
                    # Set appropriate size for logo
                    target_height_px = 80
                    scale = target_height_px / orig_height
                    img.height = target_height_px
                    img.width = int(orig_width * scale)
                    ws.add_image(img, addr)
                    # Don't set text value for image cells
                    continue
                except Exception as e:
                    value = f"[Logo图片错误] {value}: {e}"
            else:
                value = f"[Logo图片未找到] {value}"
        
        ws[addr] = value

    from PIL import Image as PILImage
    row = PRODUCT_START_ROW
    for product in data.get('products', []):
        # Start with standard image height
        row_height = 100  # Standard height for product rows with images
        
        for key, col in COLUMN_MAP.items():
            if key in product:
                if key == '产品图片':
                    # Use SKU (产品编号) to construct image path, check multiple directories
                    sku = product.get('产品编号', '')
                    # Try products directory first, then accessories directory
                    img_path = template.parent.parent / 'images' / 'products' / f'{sku}.jpg'
                    if not img_path.exists():
                        img_path = template.parent.parent / 'images' / 'accessories' / f'{sku}.jpg'
                    
                    if img_path.exists():
                        try:
                            # Verify image can be opened by Pillow and get size
                            with PILImage.open(img_path) as pil_img:
                                pil_img.verify()
                                orig_width, orig_height = pil_img.size
                            img = Image(str(img_path))
                            # openpyxl row height is in points (1 point = 1/72 inch),
                            # and image.height is in pixels. Excel's default DPI is 96.
                            # 1 point = 1.333 pixels, so 100 points ≈ 133 pixels
                            target_height_px = 133
                            scale = target_height_px / orig_height
                            img.height = target_height_px
                            img.width = int(orig_width * scale)
                            ws.add_image(img, f"{col}{row}")
                        except Exception as e:
                            ws[f"{col}{row}"] = f"[图片错误] {product[key]}: {e}"
                    else:
                        ws[f"{col}{row}"] = f"[图片未找到] {product[key]}"
                elif key == '描述':
                    # Handle rich text formatting for descriptions
                    rich_text_value = create_rich_text_from_json(product[key])
                    cell = ws[f"{col}{row}"]
                    cell.value = rich_text_value
                    
                    # Ensure text wrapping is enabled
                    # Create new Alignment to avoid deprecation warning
                    if cell.alignment:
                        cell.alignment = Alignment(
                            horizontal=cell.alignment.horizontal,
                            vertical=cell.alignment.vertical,
                            text_rotation=cell.alignment.text_rotation,
                            indent=cell.alignment.indent,
                            shrink_to_fit=cell.alignment.shrink_to_fit,
                            wrap_text=True
                        )
                    else:
                        cell.alignment = Alignment(wrap_text=True)
                    
                    # Calculate required height based on text length
                    # Get text content from description
                    description_text = ""
                    if isinstance(product[key], str):
                        description_text = product[key]
                    elif isinstance(product[key], dict) and product[key].get('type') == 'rich_text':
                        # Concatenate all text blocks
                        description_text = ''.join(
                            block.get('text', '') 
                            for block in product[key].get('content', [])
                        )
                    
                    # Count newlines in the text
                    newline_count = description_text.count('\n')
                    
                    # Estimate lines needed based on character count and column width
                    # Column C (描述) width is ~35.86 Excel units, actual wrapping width is narrower
                    # For Chinese/mixed text with wrapping, estimate ~30 chars per line
                    # Being conservative to ensure enough height for all text
                    chars_per_line = 30
                    text_length = len(description_text)
                    wrapped_lines = max(1, (text_length + chars_per_line - 1) // chars_per_line)
                    
                    # Total lines include both wrapped lines and explicit newlines
                    estimated_lines = wrapped_lines + newline_count
                    
                    # Each line needs about 20 points (increased for better spacing and readability)
                    # Minimum height of 50 for any description
                    estimated_height = max(50, estimated_lines * 20)
                    
                    # Only expand if description needs more space
                    if estimated_height > 100:
                        row_height = max(row_height, estimated_height)
                else:
                    ws[f"{col}{row}"] = product[key]
        
        # Set the row height
        ws.row_dimensions[row].height = row_height
        qty = product.get('数量/个')
        price = product.get('单价')
        if qty not in (None, '') and price not in (None, ''):
            ws[f"F{row}"] = f"=D{row}*E{row}"
        row += 1

    footer = data.get('footer', {})
    if 'buyer' in footer:
        ws['B69'] = footer['buyer']
    if 'supplier' in footer:
        ws['E69'] = footer['supplier']
    return wb


def main(argv: list[str]) -> int:
    if len(argv) != 3:
        print("usage: json_PO_excel.py <input.json> <output.xlsx>")
        return 1

    json_path = Path(argv[1])
    out_path = Path(argv[2])
    
    # Enforce that output must be in PO_excel_export folder
    po_excel_export_dir = Path(__file__).resolve().parent / 'PO_excel_export'
    
    # Ensure PO_excel_export directory exists
    po_excel_export_dir.mkdir(exist_ok=True)
    
    # Check if output path is trying to write to PO_excel folder
    try:
        resolved_out = out_path.resolve()
        if 'PO_excel' in str(resolved_out) and 'PO_excel_export' not in str(resolved_out):
            # User is trying to write to PO_excel folder (but not PO_excel_export)
            if resolved_out.parent.name == 'PO_excel':
                print("=" * 70)
                print("ERROR: Cannot generate files into PO_excel/ folder!")
                print("=" * 70)
                print()
                print("The PO_excel/ folder is reserved for source Excel files.")
                print("Generated files must go to PO_excel_export/ folder.")
                print()
                print("Correct usage:")
                print(f"  python json_PO_excel.py {json_path.name} PO_excel_export/{out_path.name}")
                print()
                print("=" * 70)
                return 1
    except:
        pass
    
    # Auto-correct if user provides just a filename (no directory path)
    # Check if the output path doesn't include PO_excel_export and is just a filename
    if 'PO_excel_export' not in str(out_path) and len(out_path.parts) == 1:
        # Automatically place in PO_excel_export
        out_path = po_excel_export_dir / out_path.name
        print(f"Note: Auto-placing output in PO_excel_export/")
        print(f"      {out_path}")
    
    template = Path(__file__).resolve().parent / 'docs' / 'empty_base_template.xlsx'

    with open(json_path, 'r', encoding='utf-8-sig') as f:
        data = json.load(f)

    wb = fill_workbook(template, data, json_path.name)
    wb.save(out_path)
    print(f"[OK] Generated: {out_path}")
    return 0


if __name__ == '__main__':
    raise SystemExit(main(sys.argv))
