#!/usr/bin/env python3
"""
Excel to JSON Template Generator with GUI

This script converts Excel files in the format of empty_base_template.xlsx
to JSON template files suitable for the json_template folder.

The script:
1. Reads Excel files with the standard template format
2. Extracts cell data, product information, and footer data
3. Generates individual JSON template files for each product SKU
4. Saves files to the json_template directory with proper formatting

Usage:
    python excel_to_json_template.py                    # Launch GUI (recommended)
    python excel_to_json_template.py input_file.xlsx    # Command line mode
    python excel_to_json_template.py *.xlsx             # Process multiple files

GUI Features:
- File selection with browse dialog
- Folder selection to process all Excel files
- Real-time conversion progress
- Detailed logging and error reporting
- Background processing to keep UI responsive
"""

import json
import re
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.rich_text import CellRichText
except ImportError:
    print("Error: openpyxl is required. Install it with: pip install openpyxl")
    sys.exit(1)


class ExcelToJsonConverter:
    def __init__(self):
        self.root_dir = Path(__file__).resolve().parent
        self.template_dir = self.root_dir / "json_template"
        self.images_dir = self.root_dir / "images"
        
        # Create template directory if it doesn't exist
        self.template_dir.mkdir(exist_ok=True)
        
        # Load accessory mapping for product names
        self.accessory_map = self._load_accessory_mapping()
        
    def _load_accessory_mapping(self) -> Dict[str, Dict]:
        """Load accessory mapping to get product names"""
        mapping_path = self.root_dir / "docs" / "accessory_mapping.json"
        try:
            with open(mapping_path, "r", encoding="utf-8-sig") as f:
                data = json.load(f)
                return data.get("products", {})
        except FileNotFoundError:
            print(f"Warning: {mapping_path} not found. Product names may not be populated.")
            return {}
    
    def _find_image_path(self, sku: str) -> Optional[str]:
        """Find image path for the given SKU"""
        for sub in ("products", "accessories"):
            img_path = self.images_dir / sub / f"{sku}.jpg"
            if img_path.exists():
                return f"order_generation/images/{sub}/{sku}.jpg"
        return None
    
    def _get_cell_value(self, ws, row: int, col: int) -> str:
        """Get cell value as string, handling formulas and rich text"""
        cell = ws.cell(row=row, column=col)
        
        # Handle different cell value types
        if cell.value is None:
            return ""
        elif isinstance(cell.value, str):
            return cell.value.strip()
        elif isinstance(cell.value, CellRichText):
            # Convert CellRichText to plain text for non-description fields
            return str(cell.value).strip()
        else:
            return str(cell.value).strip()
    
    def _extract_color_from_excel_font(self, font_color):
        """
        Extract hex color from Excel font color object.
        Handles RGB, indexed, and theme colors with proper fallbacks.
        """
        if not font_color:
            return "000000"
        
        # Handle RGB colors (most common)
        if hasattr(font_color, 'rgb') and font_color.rgb:
            color_val = str(font_color.rgb)
            # Remove alpha channel if present (first 2 characters)
            if len(color_val) == 8:
                return color_val[2:]  # Remove alpha channel
            elif len(color_val) == 6:
                return color_val
        
        # Handle indexed colors (Excel's predefined palette)
        if hasattr(font_color, 'indexed') and font_color.indexed is not None:
            indexed_colors = {
                0: "000000",  # Black
                1: "FFFFFF",  # White  
                2: "FF0000",  # Red
                3: "00FF00",  # Green
                4: "0000FF",  # Blue
                5: "FFFF00",  # Yellow
                6: "FF00FF",  # Magenta
                7: "00FFFF",  # Cyan
                8: "000000",  # Black (duplicate)
                9: "FFFFFF",  # White (duplicate)
                10: "FF0000", # Red (duplicate)
                11: "00FF00", # Green (duplicate)  
                12: "0000FF", # Blue (duplicate)
                13: "FFFF00", # Yellow (duplicate)
                14: "FF00FF", # Magenta (duplicate)
                15: "00FFFF", # Cyan (duplicate)
                16: "800000", # Dark Red
                17: "008000", # Dark Green
                18: "000080", # Dark Blue
                19: "808000", # Dark Yellow
                20: "800080", # Dark Magenta
                21: "008080", # Dark Cyan
                22: "C0C0C0", # Light Gray
                23: "808080", # Gray
            }
            return indexed_colors.get(font_color.indexed, "000000")
        
        # Handle theme colors (convert to reasonable defaults)
        if hasattr(font_color, 'theme') and font_color.theme is not None:
            # Excel theme colors - map to common equivalents
            theme_colors = {
                0: "FFFFFF",  # Background 1 (Light)
                1: "000000",  # Text 1 (Dark)
                2: "FFFFFF",  # Background 2 (Light)
                3: "000000",  # Text 2 (Dark)  
                4: "0070C0",  # Accent 1 (Blue)
                5: "FF0000",  # Accent 2 (Red)
                6: "00B050",  # Accent 3 (Green)
                7: "7030A0",  # Accent 4 (Purple)
                8: "0099CC",  # Accent 5 (Light Blue)
                9: "FF9900",  # Accent 6 (Orange)
            }
            base_color = theme_colors.get(font_color.theme, "000000")
            
            # Apply tint if present
            if hasattr(font_color, 'tint') and font_color.tint != 0:
                # Simplified tint application - just return base color
                # (Full tint calculation is complex and rarely needed)
                return base_color
            
            return base_color
        
        # Fallback to black for unknown color types
        return "000000"
    
    def _parse_text_formatting_tags(self, text: str):
        """
        Parse text formatting tags like [BOLD:RED]text[/BOLD] 
        Returns rich text structure or plain text if no tags found.
        """
        import re
        
        # Pattern to match [BOLD:COLOR]text[/BOLD] or [NORMAL:COLOR]text[/NORMAL]
        pattern = r'\[([^:]+):([^\]]+)\]([^\[]*)\[/[^\]]+\]'
        matches = re.findall(pattern, text)
        
        if not matches:
            return text  # No formatting tags found, return plain text
        
        rich_text_content = []
        last_end = 0
        
        # Find all matches with their positions
        for match in re.finditer(pattern, text):
            # Add any text before this match as plain text
            if match.start() > last_end:
                plain_part = text[last_end:match.start()]
                if plain_part:
                    rich_text_content.append({
                        "text": plain_part,
                        "bold": False,
                        "color": "000000"
                    })
            
            # Parse the formatted part
            format_type = match.group(1).upper()
            color = match.group(2).upper()
            formatted_text = match.group(3)
            
            # Convert color names to hex (expanded list)
            color_map = {
                # Basic colors
                'RED': 'FF0000',
                'GREEN': '008000',
                'BLUE': '0000FF', 
                'BLACK': '000000',
                'WHITE': 'FFFFFF',
                'YELLOW': 'FFFF00',
                'PURPLE': '800080',
                'ORANGE': 'FFA500',
                
                # Extended colors
                'DARKRED': '800000',
                'DARKGREEN': '006400',
                'DARKBLUE': '000080',
                'LIGHTBLUE': '87CEEB',
                'LIGHTGREEN': '90EE90',
                'PINK': 'FFC0CB',
                'BROWN': 'A52A2A',
                'GRAY': '808080',
                'GREY': '808080',
                'LIGHTGRAY': 'D3D3D3',
                'LIGHTGREY': 'D3D3D3',
                'DARKGRAY': 'A9A9A9',
                'DARKGREY': 'A9A9A9',
                'CYAN': '00FFFF',
                'MAGENTA': 'FF00FF',
                'LIME': '00FF00',
                'MAROON': '800000',
                'NAVY': '000080',
                'OLIVE': '808000',
                'SILVER': 'C0C0C0',
                'TEAL': '008080',
                'AQUA': '00FFFF',
                'FUCHSIA': 'FF00FF',
            }
            
            hex_color = color_map.get(color, color)
            
            # Validate hex color format (6 hex digits)
            if re.match(r'^[0-9A-F]{6}$', hex_color):
                validated_color = hex_color
            elif re.match(r'^[0-9A-F]{8}$', hex_color):
                # Remove alpha channel if 8 digits
                validated_color = hex_color[2:]
            else:
                # Invalid format, default to black
                validated_color = '000000'
                print(f"Warning: Invalid color '{color}' in text tag, using black")
            
            rich_text_content.append({
                "text": formatted_text,
                "bold": format_type == 'BOLD',
                "color": validated_color
            })
            
            last_end = match.end()
        
        # Add any remaining text after the last match
        if last_end < len(text):
            remaining = text[last_end:]
            if remaining:
                rich_text_content.append({
                    "text": remaining,
                    "bold": False,
                    "color": "000000"
                })
        
        # If we have rich content, return structured format
        if rich_text_content and len(rich_text_content) > 1 or (len(rich_text_content) == 1 and rich_text_content[0]["bold"]):
            return {
                "type": "rich_text",
                "content": rich_text_content
            }
        
        return text  # Return original if no meaningful formatting
    
    def _extract_rich_text_from_cell(self, ws, row: int, col: int):
        """
        Extract rich text formatting from Excel cell.
        Handles multiple scenarios:
        1. CellRichText objects (ideal case)
        2. Cell with mixed formatting (requires special handling)
        3. Plain text with text-based formatting tags
        4. Plain text fallback
        """
        cell = ws.cell(row=row, column=col)
        
        # Check if cell is empty
        if not cell.value:
            return {
                "type": "rich_text",
                "content": [{
                    "text": "",
                    "bold": False,
                    "color": "000000"
                }]
            }
        
        # SCENARIO 1: Check if cell contains rich text (CellRichText object)
        if isinstance(cell.value, CellRichText):
            return self._extract_cellrichtext_formatting(cell.value)
        
        # SCENARIO 2: Check for mixed formatting within cell using alternative methods
        rich_text_result = self._extract_mixed_formatting(ws, row, col)
        if rich_text_result:
            return rich_text_result
        
        # SCENARIO 3: Regular cell - check for basic formatting or text tags
        text = str(cell.value).strip()
        
        # Check for text-based formatting tags first [BOLD:COLOR]text[/BOLD]
        formatted_result = self._parse_text_formatting_tags(text)
        if isinstance(formatted_result, dict):
            return formatted_result
        
        # SCENARIO 4: Check if cell has any basic formatting applied
        font = cell.font
        if font and (font.bold or (font.color and self._has_significant_color(font.color))):
            # Cell has formatting - create rich text structure
            bold = bool(font.bold) if font.bold is not None else False
            color = self._extract_color_from_excel_font(font.color)
            
            return {
                "type": "rich_text", 
                "content": [{
                    "text": text,
                    "bold": bold,
                    "color": color
                }]
            }
        
        # SCENARIO 5: Convert plain text to uniform rich text format
        return {
            "type": "rich_text",
            "content": [{
                "text": text,
                "bold": False,
                "color": "000000"
            }]
        }
    
    def _extract_cellrichtext_formatting(self, cell_rich_text):
        """Extract formatting from CellRichText object"""
        rich_text_content = []
        
        for text_block in cell_rich_text:
            # Get text content
            if hasattr(text_block, 'text'):
                text = str(text_block.text)
            else:
                text = str(text_block)
            
            # Get formatting properties
            bold = False
            color = "000000"
            
            if hasattr(text_block, 'font') and text_block.font is not None:
                font = text_block.font
                
                # Check for bold
                if hasattr(font, 'b') and font.b is not None:
                    bold = bool(font.b)
                
                # Check for color
                color = self._extract_color_from_excel_font(font.color)
            
            rich_text_content.append({
                "text": text,
                "bold": bold,
                "color": color
            })
        
        return {
            "type": "rich_text",
            "content": rich_text_content
        }
    
    def _extract_mixed_formatting(self, ws, row: int, col: int):
        """
        Try to extract mixed formatting from cells that don't use CellRichText.
        This method directly parses Excel XML to extract actual rich text formatting.
        """
        try:
            cell = ws.cell(row=row, column=col)
            
            # Method 1: Direct XML parsing for shared strings with rich text
            xml_result = self._extract_rich_text_from_xml(ws, row, col)
            if xml_result:
                return xml_result
            
            # Method 2: Check openpyxl shared strings (fallback)
            if hasattr(ws.parent, 'shared_strings') and ws.parent.shared_strings:
                cell_value = str(cell.value)
                for shared_string in ws.parent.shared_strings:
                    if str(shared_string) == cell_value:
                        if hasattr(shared_string, 'runs') and shared_string.runs:
                            return self._parse_shared_string_runs(shared_string.runs)
            
            # No intelligent formatting - only use actual Excel formatting
            
        except Exception as e:
            print(f"Debug: Advanced formatting extraction failed for cell {row},{col}: {e}")
        
        return None
    
    def _extract_rich_text_from_xml(self, ws, row: int, col: int):
        """
        Extract rich text formatting by directly parsing Excel XML.
        This handles cases where openpyxl doesn't properly read shared string rich text.
        """
        try:
            import zipfile
            import xml.etree.ElementTree as ET
            
            # Get the workbook's file path - try different approaches
            excel_file_path = None
            
            # Method 1: Check our custom stored path
            if hasattr(ws.parent, '_excel_file_path'):
                excel_file_path = ws.parent._excel_file_path
            
            # Method 2: Check if workbook has path attribute
            elif hasattr(ws.parent, 'path'):
                excel_file_path = ws.parent.path
            
            # Method 3: Check _archive attribute (older openpyxl versions)
            elif hasattr(ws.parent, '_archive') and hasattr(ws.parent._archive, 'filename'):
                excel_file_path = ws.parent._archive.filename
            
            if not excel_file_path:
                return None
            
            cell = ws.cell(row=row, column=col)
            
            with zipfile.ZipFile(excel_file_path, 'r') as zip_file:
                # Get worksheet index (1-based for Excel)
                worksheet_index = 1
                if hasattr(ws.parent, 'worksheets'):
                    try:
                        worksheet_index = ws.parent.worksheets.index(ws) + 1
                    except:
                        worksheet_index = 1
                
                # Read worksheet XML
                worksheet_file = f'xl/worksheets/sheet{worksheet_index}.xml'
                if worksheet_file not in zip_file.namelist():
                    return None
                    
                sheet_xml = zip_file.read(worksheet_file).decode('utf-8')
                
                # Parse worksheet XML to find the cell
                sheet_root = ET.fromstring(sheet_xml)
                namespaces = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                
                # Convert row,col to Excel reference (e.g., 7,3 -> C7)
                cell_ref = f"{chr(64 + col)}{row}"
                
                # Find the specific cell in XML
                cells = sheet_root.findall('.//main:c', namespaces)
                
                shared_string_index = None
                for xml_cell in cells:
                    if xml_cell.get('r') == cell_ref and xml_cell.get('t') == 's':
                        value_elem = xml_cell.find('main:v', namespaces)
                        if value_elem is not None:
                            shared_string_index = int(value_elem.text)
                            break
                
                if shared_string_index is None:
                    return None
                
                # Read shared strings XML
                if 'xl/sharedStrings.xml' not in zip_file.namelist():
                    return None
                    
                ss_xml = zip_file.read('xl/sharedStrings.xml').decode('utf-8')
                ss_root = ET.fromstring(ss_xml)
                
                # Find the specific shared string
                si_elements = ss_root.findall('.//main:si', namespaces)
                if len(si_elements) <= shared_string_index:
                    return None
                
                target_si = si_elements[shared_string_index]
                
                # Extract rich text runs
                runs = target_si.findall('main:r', namespaces)
                if not runs:
                    return None
                
                rich_text_segments = []
                for run in runs:
                    # Get text content
                    text_elem = run.find('main:t', namespaces)
                    text = text_elem.text if text_elem is not None else ""
                    
                    # Get run properties (formatting)
                    rpr = run.find('main:rPr', namespaces)
                    bold = False
                    color = "000000"
                    
                    if rpr is not None:
                        # Check for bold
                        bold_elem = rpr.find('main:b', namespaces)
                        bold = bold_elem is not None
                        
                        # Check for color
                        color_elem = rpr.find('main:color', namespaces)
                        if color_elem is not None:
                            if 'rgb' in color_elem.attrib:
                                color = color_elem.get('rgb')
                                if len(color) == 8:  # Remove alpha if present
                                    color = color[2:]
                            elif 'indexed' in color_elem.attrib:
                                indexed = int(color_elem.get('indexed'))
                                # Common indexed colors
                                indexed_colors = {
                                    10: "FF0000",  # Red
                                    9: "FFFFFF",   # White
                                    8: "000000",   # Black
                                    64: "000000",  # Auto/Black
                                }
                                color = indexed_colors.get(indexed, "000000")
                            elif 'theme' in color_elem.attrib:
                                theme = int(color_elem.get('theme'))
                                theme_colors = {
                                    0: "000000",  # Dark
                                    1: "FFFFFF",  # Light
                                    2: "FF0000",  # Red
                                }
                                color = theme_colors.get(theme, "000000")
                    
                    rich_text_segments.append({
                        "text": text,
                        "bold": bold,
                        "color": color
                    })
                
                # Return rich text structure if we found formatted content
                if rich_text_segments:
                    print(f"Debug: Successfully extracted {len(rich_text_segments)} rich text segments from XML")
                    return {
                        "type": "rich_text",
                        "content": rich_text_segments
                    }
                
        except Exception as e:
            # If XML parsing fails, continue with other methods
            print(f"Debug: XML rich text extraction failed for cell {row},{col}: {e}")
        
        return None
    
    def _parse_shared_string_runs(self, runs):
        """Parse shared string runs for rich text formatting"""
        rich_text_content = []
        
        for run in runs:
            text = getattr(run, 'text', str(run))
            
            # Get formatting from run properties
            bold = False
            color = "000000"
            
            if hasattr(run, 'font'):
                font = run.font
                if hasattr(font, 'b') and font.b:
                    bold = True
                if hasattr(font, 'color'):
                    color = self._extract_color_from_excel_font(font.color)
            
            rich_text_content.append({
                "text": text,
                "bold": bold,
                "color": color
            })
        
        if rich_text_content:
            return {
                "type": "rich_text",
                "content": rich_text_content
            }
        
        return None
    
    def _has_significant_color(self, color_obj):
        """Check if color is significantly different from default black"""
        if not color_obj:
            return False
        
        # Extract color and check if it's not default black
        extracted_color = self._extract_color_from_excel_font(color_obj)
        return extracted_color != "000000"
    
    def _extract_cells_data(self, ws) -> Dict[str, Dict[str, str]]:
        """Extract all cell data with keys and values"""
        cells = {}
        
        # Define the cell mapping based on the template structure
        # These are the standard cells that contain metadata
        cell_mappings = {
            # Row 3
            "B3": {"key": "供货商：", "row": 3, "col": 2},
            "G3": {"key": "订单号", "row": 3, "col": 7},
            
            # Row 4
            "B4": {"key": "电话：", "row": 4, "col": 2},
            "G4": {"key": "日期", "row": 4, "col": 7},
            
            # Row 5
            "B5": {"key": "联系人：", "row": 5, "col": 2},
            "G5": {"key": "订单安排人", "row": 5, "col": 7},
            
            # Row 12-17 (various fields)
            "B12": {"key": "进仓地址：", "row": 12, "col": 2},
            "B13": {"key": "付款方式", "row": 13, "col": 2},
            "B14": {"key": "交货时间", "row": 14, "col": 2},
            "F14": {"key": "色卡", "row": 14, "col": 6},
            "G14": {"key": "Logo", "row": 14, "col": 7},
            "B15": {"key": "箱规", "row": 15, "col": 2},
            "F15": {"key": "色卡1", "row": 15, "col": 6},
            "G15": {"key": "Logo1", "row": 15, "col": 7},
            "B16": {"key": "产前确认样", "row": 16, "col": 2},
            "F16": {"key": "色卡2", "row": 16, "col": 6},
            "G16": {"key": "Logo2", "row": 16, "col": 7},
            "B17": {"key": "出货样", "row": 17, "col": 2},
            "F17": {"key": "色卡3", "row": 17, "col": 6},
            "G17": {"key": "Logo3", "row": 17, "col": 7},
            "F18": {"key": "色卡4", "row": 18, "col": 6},
            "G18": {"key": "Logo4", "row": 18, "col": 7},
        }
        
        # Add note fields (A19-A30)
        for i in range(19, 31):
            if i == 19:
                key = "注意事项：1"
            else:
                key = f"{i-18}："
            cell_mappings[f"A{i}"] = {"key": key, "row": i, "col": 1}
        
        # Extract values for mapped cells
        for addr, info in cell_mappings.items():
            value = self._get_cell_value(ws, info["row"], info["col"])
            cells[addr] = {
                "key": info["key"],
                "value": value
            }
        
        return cells
    
    def _find_product_table_start(self, ws) -> int:
        """Find the row where the product table starts"""
        # Look for the header row containing "产品编号", "数量/个", etc.
        for row in range(1, 20):  # Check first 20 rows
            for col in range(1, 8):  # Check columns A-G
                cell_value = self._get_cell_value(ws, row, col)
                if cell_value in ("产品编号", "型号"):
                    return row
        return 7  # Default to row 7 if not found
    
    def _extract_products(self, ws) -> List[Dict[str, Any]]:
        """Extract product data from the worksheet"""
        products = []
        header_row = self._find_product_table_start(ws)
        
        # Define column mapping for product table
        # Based on standard template: A=产品编号, B=产品图片, C=描述, D=数量/个, E=单价, G=包装方式
        col_mapping = {
            1: "产品编号",     # Column A
            2: "产品图片",     # Column B  
            3: "描述",         # Column C
            4: "数量/个",      # Column D
            5: "单价",         # Column E
            7: "包装方式"      # Column G
        }
        
        # Start from the row after header
        row = header_row + 1
        
        while row <= ws.max_row:
            # Get product code from column A
            sku = self._get_cell_value(ws, row, 1)
            
            # Stop if we hit an empty SKU or total row
            if not sku or sku.upper().startswith("TOTAL") or sku == "总计":
                break
            
            product = {}
            
            # Extract data for each column
            for col, field in col_mapping.items():
                if field == "描述":
                    # Use rich text extraction for descriptions
                    value = self._extract_rich_text_from_cell(ws, row, col)
                else:
                    # Use regular cell value extraction for other fields
                    value = self._get_cell_value(ws, row, col)
                
                if field == "产品编号":
                    product[field] = value
                elif field == "产品图片":
                    # Try to find image path, use provided value as fallback
                    img_path = self._find_image_path(sku)
                    product[field] = img_path or value
                elif field == "数量/个":
                    # Convert to integer, default to 0
                    try:
                        product[field] = int(float(value)) if value else 0
                    except ValueError:
                        product[field] = 0
                elif field == "单价":
                    # Convert to float
                    try:
                        product[field] = float(value) if value else 0.0
                    except ValueError:
                        product[field] = 0.0
                else:
                    product[field] = value
            
            # Add product name from accessory mapping if available
            if sku in self.accessory_map:
                product["产品名称"] = self.accessory_map[sku]["name"]
            elif "产品名称" not in product:
                product["产品名称"] = ""  # Default empty name
            
            products.append(product)
            row += 1
        
        return products
    
    def _extract_footer(self, ws) -> Dict[str, str]:
        """Extract footer information (buyer, supplier)"""
        footer = {}
        
        # Look for buyer and supplier info around row 69 (standard template)
        try:
            buyer = self._get_cell_value(ws, 69, 2)  # B69
            supplier = self._get_cell_value(ws, 69, 5)  # E69
            
            if buyer:
                footer["buyer"] = buyer
            if supplier:
                footer["supplier"] = supplier
        except:
            pass
        
        return footer
    
    def convert_excel_to_json(self, excel_path: Path) -> List[Path]:
        """Convert Excel file to JSON template(s)"""
        print(f"Processing: {excel_path}")
        
        try:
            # Load without data_only=True to preserve rich text formatting
            wb = load_workbook(excel_path, data_only=False)
            
            # Store the file path for XML extraction
            wb._excel_file_path = str(excel_path)
            
            ws = wb.active
            
            # Extract data
            cells = self._extract_cells_data(ws)
            products = self._extract_products(ws)
            footer = self._extract_footer(ws)
            
            if not products:
                print(f"Warning: No products found in {excel_path}")
                return []
            
            generated_files = []
            
            # Group products by SKU to avoid duplicates
            products_by_sku = {}
            for product in products:
                sku = product.get("产品编号")
                if sku:
                    if sku not in products_by_sku:
                        products_by_sku[sku] = product
                    else:
                        # If duplicate SKU, combine quantities
                        existing = products_by_sku[sku]
                        existing["数量/个"] += product.get("数量/个", 0)
            
            # Generate one JSON file per unique product SKU
            for sku, product in products_by_sku.items():
                # Create JSON structure
                json_data = {
                    "cells": cells,
                    "products": [product],
                    "footer": footer
                }
                
                # Save to json_template directory
                output_path = self.template_dir / f"{sku}.json"
                
                with open(output_path, "w", encoding="utf-8") as f:
                    json.dump(json_data, f, ensure_ascii=False, indent=2)
                
                generated_files.append(output_path)
                print(f"  Generated: {output_path}")
            
            return generated_files
            
        except Exception as e:
            print(f"Error processing {excel_path}: {e}")
            return []
    
    def process_files(self, file_patterns: List[str]) -> None:
        """Process multiple Excel files"""
        total_generated = 0
        
        for pattern in file_patterns:
            # Handle both specific files and glob patterns
            if "*" in pattern:
                files = list(Path(".").glob(pattern))
            else:
                files = [Path(pattern)]
            
            for file_path in files:
                if file_path.suffix.lower() in (".xlsx", ".xls"):
                    generated = self.convert_excel_to_json(file_path)
                    total_generated += len(generated)
        
        print(f"\nTotal JSON templates generated: {total_generated}")
        print(f"Output directory: {self.template_dir}")


class ExcelToJsonGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel转JSON模板转换器")
        self.root.geometry("800x600")
        
        # Initialize converter
        self.converter = ExcelToJsonConverter()
        
        # Selected files list
        self.selected_files = []
        
        # Create GUI
        self._create_widgets()
        
    def _create_widgets(self):
        """Create all GUI widgets"""
        # Main frame
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(2, weight=1)
        
        # Title
        title_label = ttk.Label(main_frame, text="Excel转JSON模板转换器", 
                               font=("TkDefaultFont", 16, "bold"))
        title_label.grid(row=0, column=0, pady=(0, 20))
        
        # File selection section
        file_frame = ttk.LabelFrame(main_frame, text="文件选择", padding="10")
        file_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        file_frame.columnconfigure(0, weight=1)
        
        # File selection buttons
        button_frame = ttk.Frame(file_frame)
        button_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        ttk.Button(button_frame, text="选择Excel文件", 
                  command=self._select_files).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="添加文件夹", 
                  command=self._select_folder).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="清空全部", 
                  command=self._clear_files).pack(side=tk.LEFT, padx=(0, 10))
        
        # File list
        list_frame = ttk.Frame(file_frame)
        list_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        list_frame.columnconfigure(0, weight=1)
        list_frame.rowconfigure(0, weight=1)
        
        # Create listbox with scrollbar
        self.file_listbox = tk.Listbox(list_frame, height=6)
        file_scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.file_listbox.yview)
        self.file_listbox.configure(yscrollcommand=file_scrollbar.set)
        
        self.file_listbox.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        file_scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        # Control buttons
        control_frame = ttk.Frame(file_frame)
        control_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(10, 0))
        
        ttk.Button(control_frame, text="移除选中", 
                  command=self._remove_selected).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(control_frame, text="转换为JSON", 
                  command=self._convert_files).pack(side=tk.LEFT, padx=(20, 0))
        
        # Progress and output section
        output_frame = ttk.LabelFrame(main_frame, text="转换进度", padding="10")
        output_frame.grid(row=2, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        output_frame.columnconfigure(0, weight=1)
        output_frame.rowconfigure(1, weight=1)
        
        # Progress bar
        self.progress_var = tk.StringVar(value="准备转换文件...")
        progress_label = ttk.Label(output_frame, textvariable=self.progress_var)
        progress_label.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 5))
        
        self.progress_bar = ttk.Progressbar(output_frame, mode='determinate')
        self.progress_bar.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # Output text
        self.output_text = scrolledtext.ScrolledText(output_frame, height=15, wrap=tk.WORD)
        self.output_text.grid(row=2, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Status bar
        self.status_var = tk.StringVar(value="就绪")
        status_bar = ttk.Label(main_frame, textvariable=self.status_var, relief=tk.SUNKEN)
        status_bar.grid(row=3, column=0, sticky=(tk.W, tk.E), pady=(10, 0))
        
        # Initial message
        self._log("Excel转JSON模板转换器准备就绪！")
        self._log("选择Excel文件以转换为JSON模板。")
        self._log(f"输出目录: {self.converter.template_dir}")
        
    def _log(self, message: str):
        """Log a message to the output text"""
        self.output_text.insert(tk.END, message + "\n")
        self.output_text.see(tk.END)
        self.root.update()
        
    def _select_files(self):
        """Select Excel files to convert"""
        filetypes = [
            ("Excel文件", "*.xlsx *.xls"),
            ("所有文件", "*.*")
        ]
        
        files = filedialog.askopenfilenames(
            title="选择要转换的Excel文件",
            filetypes=filetypes
        )
        
        if files:
            added_count = 0
            for file_path in files:
                if file_path not in self.selected_files:
                    self.selected_files.append(file_path)
                    self.file_listbox.insert(tk.END, Path(file_path).name)
                    added_count += 1
            
            self._log(f"添加了 {added_count} 个文件")
            self.status_var.set(f"已选择 {len(self.selected_files)} 个文件")
    
    def _select_folder(self):
        """Select all Excel files from a folder"""
        folder = filedialog.askdirectory(title="选择包含Excel文件的文件夹")
        
        if folder:
            folder_path = Path(folder)
            excel_files = list(folder_path.glob("*.xlsx")) + list(folder_path.glob("*.xls"))
            
            added_count = 0
            for file_path in excel_files:
                file_str = str(file_path)
                if file_str not in self.selected_files:
                    self.selected_files.append(file_str)
                    self.file_listbox.insert(tk.END, file_path.name)
                    added_count += 1
            
            self._log(f"从文件夹添加了 {added_count} 个文件: {folder}")
            self.status_var.set(f"已选择 {len(self.selected_files)} 个文件")
    
    def _clear_files(self):
        """Clear all selected files"""
        self.selected_files.clear()
        self.file_listbox.delete(0, tk.END)
        self._log("清空了所有选中的文件")
        self.status_var.set("就绪")
    
    def _remove_selected(self):
        """Remove selected file from list"""
        selection = self.file_listbox.curselection()
        if selection:
            index = selection[0]
            removed_file = self.selected_files.pop(index)
            self.file_listbox.delete(index)
            self._log(f"移除了: {Path(removed_file).name}")
            self.status_var.set(f"已选择 {len(self.selected_files)} 个文件")
    
    def _convert_files(self):
        """Convert selected files to JSON templates"""
        if not self.selected_files:
            messagebox.showwarning("警告", "请选择要转换的Excel文件")
            return
        
        # Disable convert button during processing
        for child in self.root.winfo_children():
            self._disable_widgets(child)
        
        # Start conversion in background thread
        thread = threading.Thread(target=self._conversion_worker)
        thread.daemon = True
        thread.start()
    
    def _disable_widgets(self, widget):
        """Recursively disable all widgets"""
        try:
            widget.configure(state='disabled')
        except:
            pass
        for child in widget.winfo_children():
            self._disable_widgets(child)
    
    def _enable_widgets(self, widget):
        """Recursively enable all widgets"""
        try:
            widget.configure(state='normal')
        except:
            pass
        for child in widget.winfo_children():
            self._enable_widgets(child)
    
    def _conversion_worker(self):
        """Background worker for file conversion"""
        try:
            total_files = len(self.selected_files)
            total_generated = 0
            
            self.progress_bar.configure(maximum=total_files)
            
            for i, file_path in enumerate(self.selected_files):
                self.progress_var.set(f"处理中 {i+1}/{total_files}: {Path(file_path).name}")
                self.progress_bar.configure(value=i)
                
                self._log(f"\n[{i+1}/{total_files}] 处理中: {Path(file_path).name}")
                
                try:
                    generated_files = self.converter.convert_excel_to_json(Path(file_path))
                    total_generated += len(generated_files)
                    
                    if generated_files:
                        self._log(f"  ✓ 生成了 {len(generated_files)} 个JSON模板")
                        for json_file in generated_files:
                            self._log(f"    - {json_file.name}")
                    else:
                        self._log(f"  ⚠ 未生成模板（未找到产品）")
                        
                except Exception as e:
                    self._log(f"  ✗ 错误: {e}")
            
            self.progress_bar.configure(value=total_files)
            self.progress_var.set("转换完成！")
            
            # Summary
            self._log(f"\n" + "="*50)
            self._log(f"转换总结")
            self._log(f"="*50)
            self._log(f"处理的文件: {total_files}")
            self._log(f"生成的JSON模板: {total_generated}")
            self._log(f"输出目录: {self.converter.template_dir}")
            self._log(f"="*50)
            
            if total_generated > 0:
                messagebox.showinfo("成功", 
                    f"转换完成！\n\n"
                    f"处理的文件: {total_files}\n"
                    f"生成的JSON模板: {total_generated}\n"
                    f"输出目录: {self.converter.template_dir}")
            else:
                messagebox.showwarning("警告", 
                    f"转换完成但未生成JSON模板。\n"
                    f"请检查Excel文件是否包含有效的产品数据。")
            
            self.status_var.set(f"完成: 生成了 {total_generated} 个模板")
            
        except Exception as e:
            self._log(f"\n意外错误: {e}")
            messagebox.showerror("错误", f"转换失败: {e}")
            self.status_var.set("发生错误")
        
        finally:
            # Re-enable widgets
            for child in self.root.winfo_children():
                self._enable_widgets(child)


def main():
    # Check if GUI mode should be used
    if len(sys.argv) == 1:
        # No command line arguments - launch GUI
        try:
            root = tk.Tk()
            app = ExcelToJsonGUI(root)
            root.mainloop()
        except Exception as e:
            print(f"GUI Error: {e}")
            print("Falling back to command line mode...")
            print("Usage: python excel_to_json_template.py <excel_file1> [excel_file2] ...")
            sys.exit(1)
    else:
        # Command line arguments provided - use CLI mode
        print("Usage: python excel_to_json_template.py <excel_file1> [excel_file2] ...")
        print("       python excel_to_json_template.py *.xlsx")
        print("       python excel_to_json_template.py  # Launch GUI")
        
        converter = ExcelToJsonConverter()
        converter.process_files(sys.argv[1:])


if __name__ == "__main__":
    main()
