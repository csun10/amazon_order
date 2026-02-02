"""
Buyer Mapping Module

Maps SKUs to their respective buyers based on listing data.
Distinguishes between parent products and accessories.
"""

from pathlib import Path
from typing import Dict, Optional
import json

try:
    from openpyxl import load_workbook
except ImportError:
    print("Warning: openpyxl not available, buyer mapping will not work")
    load_workbook = None


# Buyer constants
BUYER_JIXIU = "宁波集秀美容科技有限公司"
BUYER_PINXIU = "宁波品秀美容科技有限公司"

# Brand to buyer mapping
BRAND_TO_BUYER = {
    "JIXIUBeauty-US": BUYER_JIXIU,
    "PinxiuBeautyUS-US-US-US": BUYER_PINXIU,
}


class BuyerMapper:
    """Manages SKU to buyer mappings"""
    
    def __init__(self):
        self.sku_to_buyer: Dict[str, str] = {}
        self.parent_skus: set = set()
        self._load_mappings()
    
    def _load_mappings(self):
        """Load SKU to buyer mappings from listing file and accessory mapping"""
        root_dir = Path(__file__).resolve().parent
        
        # Load listing file
        listing_file = root_dir / "docs" / "Listing20260202-876789694451576832.xlsx"
        if listing_file.exists() and load_workbook:
            self._load_from_listing(listing_file)
        else:
            print(f"Warning: Listing file not found: {listing_file}")
        
        # Load parent products from accessory mapping
        accessory_file = root_dir / "docs" / "accessory_mapping.json"
        if accessory_file.exists():
            self._load_parent_products(accessory_file)
        else:
            print(f"Warning: Accessory mapping file not found: {accessory_file}")
    
    def _load_from_listing(self, listing_file: Path):
        """Load SKU to buyer mappings from listing Excel file"""
        try:
            wb = load_workbook(listing_file, data_only=True)
            ws = wb.active
            
            # Column indices (1-based): Brand=6, SKU=9
            brand_col = 6
            sku_col = 9
            
            for row in range(2, ws.max_row + 1):
                brand = ws.cell(row, brand_col).value
                sku = ws.cell(row, sku_col).value
                
                if brand and sku and brand in BRAND_TO_BUYER:
                    self.sku_to_buyer[sku] = BRAND_TO_BUYER[brand]
            
            wb.close()
            print(f"Loaded buyer mappings for {len(self.sku_to_buyer)} SKUs")
            
        except Exception as e:
            print(f"Error loading listing file: {e}")
    
    def _load_parent_products(self, accessory_file: Path):
        """Load parent product SKUs from accessory mapping"""
        try:
            with open(accessory_file, 'r', encoding='utf-8-sig') as f:
                data = json.load(f)
            
            # Parent products are the keys in the "products" dict
            if "products" in data:
                self.parent_skus = set(data["products"].keys())
                print(f"Loaded {len(self.parent_skus)} parent product SKUs")
        
        except Exception as e:
            print(f"Error loading accessory mapping: {e}")
    
    def get_buyer(self, sku: str) -> Optional[str]:
        """
        Get buyer for a given SKU.
        Returns None if SKU is not a parent product or not in mapping.
        
        Args:
            sku: Product SKU
            
        Returns:
            Buyer name or None
        """
        # Only return buyer for parent products
        if sku not in self.parent_skus:
            return None
        
        return self.sku_to_buyer.get(sku)
    
    def is_parent_product(self, sku: str) -> bool:
        """Check if SKU is a parent product"""
        return sku in self.parent_skus
    
    def get_buyer_or_default(self, sku: str, default: str = BUYER_PINXIU) -> str:
        """
        Get buyer for SKU, or return default if not found.
        
        Args:
            sku: Product SKU
            default: Default buyer (defaults to BUYER_PINXIU)
            
        Returns:
            Buyer name
        """
        buyer = self.get_buyer(sku)
        return buyer if buyer else default


# Global instance for easy import
_buyer_mapper = None

def get_buyer_mapper() -> BuyerMapper:
    """Get or create the global buyer mapper instance"""
    global _buyer_mapper
    if _buyer_mapper is None:
        _buyer_mapper = BuyerMapper()
    return _buyer_mapper


# Convenience functions
def get_buyer_for_sku(sku: str) -> Optional[str]:
    """Get buyer for a given SKU (only for parent products)"""
    return get_buyer_mapper().get_buyer(sku)


def is_parent_product(sku: str) -> bool:
    """Check if SKU is a parent product"""
    return get_buyer_mapper().is_parent_product(sku)


if __name__ == "__main__":
    # Test the module
    mapper = BuyerMapper()
    
    print("\n=== Buyer Mapping Test ===")
    print(f"Total SKU mappings: {len(mapper.sku_to_buyer)}")
    print(f"Total parent products: {len(mapper.parent_skus)}")
    
    # Test some SKUs
    test_skus = ["ST1122-1", "Elasticbrush01", "B10-MJB2-BK", "48-82P3-QSFG"]
    print("\nTest SKU buyers:")
    for sku in test_skus:
        buyer = mapper.get_buyer(sku)
        is_parent = mapper.is_parent_product(sku)
        print(f"  {sku}: {buyer} (parent: {is_parent})")
