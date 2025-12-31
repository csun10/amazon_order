#!/usr/bin/env python3
"""
Accessory Mapping Updater GUI

This GUI application allows users to update the accessory_mapping.json file
using Excel files that contain parent-child product relationships and accessory information.

The application can process Excel files similar to "导出产品-按SKU-*.xlsx" that contain
a "关联辅料" (Related Accessories) tab with parent-child relationships and accessory ratios.

Features:
- Browse and select Excel files with accessory data
- Preview changes before applying
- Backup existing mapping before updating
- Validate data integrity
- Support for multiple Excel file formats
- Real-time preview of mapping changes

Usage:
    python accessory_mapping_updater_gui.py
"""

import json
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime
import traceback
import urllib.request
import urllib.parse
import urllib.error
from threading import Thread
import time

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class AccessoryMappingUpdaterGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("辅料映射更新器")
        self.root.geometry("1000x700")
        
        # Initialize variables
        self.current_mapping = {}
        self.new_mapping = {}
        self.excel_data = []
        self.selected_file = None
        
        # Load current mapping
        self.mapping_file = Path(__file__).resolve().parent / "docs" / "accessory_mapping.json"
        self._load_current_mapping()
        
        # Initialize status variable early
        self.status_var = tk.StringVar(value="Ready - Load an Excel file to begin")
        
        # Create GUI
        self._create_widgets()
        
    def _load_current_mapping(self):
        """Load the current accessory mapping"""
        try:
            if self.mapping_file.exists():
                with open(self.mapping_file, 'r', encoding='utf-8') as f:
                    self.current_mapping = json.load(f)
            else:
                self.current_mapping = {"products": {}}
        except Exception as e:
            messagebox.showerror("错误", f"加载当前映射失败: {e}")
            self.current_mapping = {"products": {}}
    
    def _create_widgets(self):
        """Create all GUI widgets"""
        # Main notebook for tabs
        notebook = ttk.Notebook(self.root)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # File Selection Tab
        file_frame = ttk.Frame(notebook)
        notebook.add(file_frame, text="File Selection")
        self._create_file_selection_tab(file_frame)
        
        # Preview Tab
        preview_frame = ttk.Frame(notebook)
        notebook.add(preview_frame, text="Preview Changes")
        self._create_preview_tab(preview_frame)
        
        # Current Mapping Tab
        current_frame = ttk.Frame(notebook)
        notebook.add(current_frame, text="Current Mapping")
        self._create_current_mapping_tab(current_frame)
        
        # Image Download Tab
        image_frame = ttk.Frame(notebook)
        notebook.add(image_frame, text="Image Download")
        self._create_image_download_tab(image_frame)
        
        # Status bar
        status_bar = ttk.Label(self.root, textvariable=self.status_var, relief=tk.SUNKEN)
        status_bar.pack(side=tk.BOTTOM, fill=tk.X, padx=10, pady=(0, 10))
    
    def _create_file_selection_tab(self, parent):
        """Create file selection and processing tab"""
        # File selection section
        file_section = ttk.LabelFrame(parent, text="Excel文件选择", padding="10")
        file_section.pack(fill=tk.X, padx=10, pady=10)
        
        # File path display
        self.file_path_var = tk.StringVar()
        file_path_frame = ttk.Frame(file_section)
        file_path_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(file_path_frame, text="选择的文件:").pack(side=tk.LEFT)
        file_path_entry = ttk.Entry(file_path_frame, textvariable=self.file_path_var, state="readonly")
        file_path_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(10, 10))
        
        ttk.Button(file_path_frame, text="浏览...", command=self._browse_file).pack(side=tk.RIGHT)
        
        # Excel processing options
        options_frame = ttk.LabelFrame(file_section, text="处理选项", padding="10")
        options_frame.pack(fill=tk.X, pady=(10, 0))
        
        # Sheet selection
        sheet_frame = ttk.Frame(options_frame)
        sheet_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(sheet_frame, text="工作表名称:").pack(side=tk.LEFT)
        self.sheet_var = tk.StringVar(value="关联辅料")
        sheet_combo = ttk.Combobox(sheet_frame, textvariable=self.sheet_var, width=20)
        sheet_combo.pack(side=tk.LEFT, padx=(10, 20))
        self.sheet_combo = sheet_combo
        
        # Process button
        ttk.Button(sheet_frame, text="处理文件", command=self._process_file).pack(side=tk.LEFT, padx=(20, 0))
        
        # Column mapping section
        mapping_section = ttk.LabelFrame(parent, text="Column Mapping", padding="10")
        mapping_section.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Column mapping frame
        col_frame = ttk.Frame(mapping_section)
        col_frame.pack(fill=tk.X)
        
        # Define expected columns and their variables
        self.column_vars = {
            "主产品SKU": tk.StringVar(value="*SKU"),
            "辅料SKU": tk.StringVar(value="辅料SKU"),
            "辅料名称": tk.StringVar(value="辅料品名"),
            "主产品数量": tk.StringVar(value="辅料比例(主料)"),
            "辅料数量": tk.StringVar(value="辅料比例(辅料)")
        }
        
        for i, (label, var) in enumerate(self.column_vars.items()):
            row_frame = ttk.Frame(col_frame)
            row_frame.pack(fill=tk.X, pady=2)
            
            ttk.Label(row_frame, text=f"{label}:", width=12).pack(side=tk.LEFT)
            combo = ttk.Combobox(row_frame, textvariable=var, width=20)
            combo.pack(side=tk.LEFT, padx=(10, 0))
            setattr(self, f"col_combo_{i}", combo)
        
        # Data preview
        preview_section = ttk.LabelFrame(mapping_section, text="Data Preview", padding="10")
        preview_section.pack(fill=tk.BOTH, expand=True, pady=(10, 0))
        
        # Create treeview for data preview
        columns = ("主产品SKU", "辅料SKU", "辅料名称", "主产品数量", "辅料数量")
        self.data_tree = ttk.Treeview(preview_section, columns=columns, show='headings', height=8)
        
        for col in columns:
            self.data_tree.heading(col, text=col)
            self.data_tree.column(col, width=120)
        
        # Scrollbar for data preview
        data_scrollbar = ttk.Scrollbar(preview_section, orient=tk.VERTICAL, command=self.data_tree.yview)
        self.data_tree.configure(yscrollcommand=data_scrollbar.set)
        
        self.data_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        data_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    
    def _create_preview_tab(self, parent):
        """Create preview changes tab"""
        # Preview controls
        control_frame = ttk.Frame(parent)
        control_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Button(control_frame, text="Generate Preview", command=self._generate_preview).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(control_frame, text="Apply Changes", command=self._apply_changes).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(control_frame, text="Create Backup", command=self._create_backup).pack(side=tk.LEFT)
        
        # Preview display
        preview_text_frame = ttk.LabelFrame(parent, text="Preview Changes", padding="10")
        preview_text_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
        
        self.preview_text = scrolledtext.ScrolledText(preview_text_frame, wrap=tk.WORD, height=20)
        self.preview_text.pack(fill=tk.BOTH, expand=True)
    
    def _create_current_mapping_tab(self, parent):
        """Create current mapping display tab"""
        # Controls
        control_frame = ttk.Frame(parent)
        control_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Button(control_frame, text="Refresh", command=self._refresh_current_mapping).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Label(control_frame, text="Search:").pack(side=tk.LEFT, padx=(20, 5))
        
        self.search_var = tk.StringVar()
        self.search_var.trace('w', self._filter_current_mapping)
        search_entry = ttk.Entry(control_frame, textvariable=self.search_var, width=30)
        search_entry.pack(side=tk.LEFT)
        
        # Current mapping display
        mapping_frame = ttk.LabelFrame(parent, text="Current Accessory Mapping", padding="10")
        mapping_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
        
        # Treeview for current mapping
        map_columns = ("Product SKU", "Product Name", "Accessory SKU", "Accessory Name", "Ratio Main", "Ratio Accessory")
        self.mapping_tree = ttk.Treeview(mapping_frame, columns=map_columns, show='headings', height=15)
        
        for col in map_columns:
            self.mapping_tree.heading(col, text=col)
            self.mapping_tree.column(col, width=150)
        
        # Scrollbar for mapping
        mapping_scrollbar = ttk.Scrollbar(mapping_frame, orient=tk.VERTICAL, command=self.mapping_tree.yview)
        self.mapping_tree.configure(yscrollcommand=mapping_scrollbar.set)
        
        self.mapping_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        mapping_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Load initial mapping display
        self._refresh_current_mapping()
    
    def _create_image_download_tab(self, parent):
        """Create image download tab"""
        # Instructions
        instructions = ttk.LabelFrame(parent, text="Instructions", padding="10")
        instructions.pack(fill=tk.X, padx=10, pady=10)
        
        instruction_text = (
            "Select Excel files containing SKU and image URL columns:\n"
            "- 产品管理-辅料列表-*.xlsx (Accessory list with images)\n"
            "- 导出产品-按SKU-*.xlsx (Product list with images)\n\n"
            "Images will be downloaded to:\n"
            "- order_generation/images/accessories/ (for accessory SKUs)\n"
            "- order_generation/images/products/ (for product SKUs)"
        )
        ttk.Label(instructions, text=instruction_text, justify=tk.LEFT).pack()
        
        # File selection
        file_section = ttk.LabelFrame(parent, text="File Selection", padding="10")
        file_section.pack(fill=tk.X, padx=10, pady=10)
        
        self.image_files_var = tk.StringVar(value="No files selected")
        self.image_files = []
        
        ttk.Label(file_section, textvariable=self.image_files_var, wraplength=800).pack(pady=5)
        
        button_frame = ttk.Frame(file_section)
        button_frame.pack(pady=5)
        ttk.Button(button_frame, text="Select Excel Files", command=self._browse_image_files).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Clear Selection", command=self._clear_image_files).pack(side=tk.LEFT, padx=5)
        
        # Download options
        options_section = ttk.LabelFrame(parent, text="Download Options", padding="10")
        options_section.pack(fill=tk.X, padx=10, pady=10)
        
        self.overwrite_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(options_section, text="Overwrite existing images", variable=self.overwrite_var).pack(anchor=tk.W)
        
        # Progress section
        progress_section = ttk.LabelFrame(parent, text="Download Progress", padding="10")
        progress_section.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        self.image_progress_var = tk.StringVar(value="Ready to download")
        ttk.Label(progress_section, textvariable=self.image_progress_var).pack(pady=5)
        
        self.image_progress_bar = ttk.Progressbar(progress_section, mode='determinate')
        self.image_progress_bar.pack(fill=tk.X, padx=5, pady=5)
        
        # Log area
        log_frame = ttk.Frame(progress_section)
        log_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        self.image_log = scrolledtext.ScrolledText(log_frame, height=15, wrap=tk.WORD)
        self.image_log.pack(fill=tk.BOTH, expand=True)
        
        # Download button
        ttk.Button(progress_section, text="Start Download", command=self._start_image_download).pack(pady=10)
    
    def _browse_image_files(self):
        """Browse for Excel files with image URLs"""
        filetypes = [("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        
        filenames = filedialog.askopenfilenames(
            title="Select Excel files with image data",
            filetypes=filetypes
        )
        
        if filenames:
            self.image_files = [Path(f) for f in filenames]
            file_list = "\n".join([f.name for f in self.image_files])
            self.image_files_var.set(f"Selected {len(self.image_files)} file(s):\n{file_list}")
            self._log_image(f"[i] Selected {len(self.image_files)} file(s)")
    
    def _clear_image_files(self):
        """Clear selected files"""
        self.image_files = []
        self.image_files_var.set("No files selected")
        self._log_image("[i] File selection cleared")
    
    def _log_image(self, message):
        """Log message to image download log"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.image_log.insert(tk.END, f"[{timestamp}] {message}\n")
        self.image_log.see(tk.END)
        self.root.update_idletasks()
    
    def _start_image_download(self):
        """Start image download in background thread"""
        if not self.image_files:
            messagebox.showwarning("Warning", "Please select Excel files first")
            return
        
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("Error", "openpyxl is required for reading Excel files")
            return
        
        # Run in separate thread to avoid blocking UI
        thread = Thread(target=self._download_images_from_excel, daemon=True)
        thread.start()
    
    def _download_images_from_excel(self):
        """Download images from Excel files"""
        try:
            self._log_image("[i] Starting image download process...")
            
            # Ensure image directories exist
            accessories_dir = self.base_path / "images" / "accessories"
            products_dir = self.base_path / "images" / "products"
            accessories_dir.mkdir(parents=True, exist_ok=True)
            products_dir.mkdir(parents=True, exist_ok=True)
            
            # Load current mapping to determine product vs accessory
            accessory_skus = set()
            if self.current_mapping and "products" in self.current_mapping:
                for product_data in self.current_mapping["products"].values():
                    if "accessories" in product_data:
                        for acc in product_data["accessories"]:
                            if "accessory_sku" in acc:
                                accessory_skus.add(acc["accessory_sku"])
            
            self._log_image(f"[i] Loaded {len(accessory_skus)} accessory SKUs from mapping")
            
            all_images = []
            
            # Parse all Excel files
            for excel_file in self.image_files:
                self._log_image(f"[i] Reading file: {excel_file.name}")
                images_from_file = self._parse_excel_for_images(excel_file)
                all_images.extend(images_from_file)
                self._log_image(f"[i] Found {len(images_from_file)} image entries in {excel_file.name}")
            
            if not all_images:
                self._log_image("[!] No images found in the selected files")
                self.image_progress_var.set("No images found")
                return
            
            # Download images
            total = len(all_images)
            self.image_progress_bar['maximum'] = total
            success_count = 0
            skip_count = 0
            error_count = 0
            
            for idx, (sku, image_url) in enumerate(all_images, 1):
                # Determine target directory
                if sku in accessory_skus:
                    target_dir = accessories_dir
                    folder_name = "accessories"
                else:
                    target_dir = products_dir
                    folder_name = "products"
                
                # Try to determine file extension from URL
                parsed_url = urllib.parse.urlparse(image_url)
                path = parsed_url.path
                ext = Path(path).suffix if path else ".jpg"
                if not ext or len(ext) > 5:
                    ext = ".jpg"  # Default to .jpg if no valid extension
                
                target_file = target_dir / f"{sku}{ext}"
                
                # Check if file exists
                if target_file.exists() and not self.overwrite_var.get():
                    self._log_image(f"[i] Skip {sku} - file already exists")
                    skip_count += 1
                else:
                    # Download image
                    try:
                        urllib.request.urlretrieve(image_url, target_file)
                        self._log_image(f"[OK] {sku} -> {folder_name}/{target_file.name}")
                        success_count += 1
                    except Exception as e:
                        self._log_image(f"[FAIL] {sku} - {str(e)}")
                        error_count += 1
                
                # Update progress
                self.image_progress_bar['value'] = idx
                self.image_progress_var.set(f"Progress: {idx}/{total} (Success: {success_count}, Skip: {skip_count}, Error: {error_count})")
                time.sleep(0.05)  # Small delay to avoid overwhelming the server
            
            # Final summary
            self._log_image("\n" + "="*50)
            self._log_image("[i] Download Complete!")
            self._log_image(f"[i] Total: {total}")
            self._log_image(f"[i] Success: {success_count}")
            self._log_image(f"[i] Skipped: {skip_count}")
            self._log_image(f"[i] Errors: {error_count}")
            self.image_progress_var.set(f"Complete - Success: {success_count}, Skip: {skip_count}, Error: {error_count}")
            
        except Exception as e:
            self._log_image(f"[FAIL] Error during download: {str(e)}")
            self._log_image(traceback.format_exc())
            self.image_progress_var.set("Error occurred")
    
    def _parse_excel_for_images(self, excel_file: Path) -> List[Tuple[str, str]]:
        """Parse Excel file to extract SKU and image URL pairs"""
        results = []
        
        try:
            wb = load_workbook(excel_file, data_only=True)
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                # Find SKU and image URL columns
                sku_col = None
                image_col = None
                
                # Check first row for headers
                for col_idx, cell in enumerate(sheet[1], 1):
                    if cell.value:
                        header = str(cell.value).strip()
                        # Look for SKU column
                        if header in ["SKU", "sku", "产品SKU", "SKU编号"]:
                            sku_col = col_idx
                        # Look for image URL column
                        elif header in ["主图", "图片", "图片链接", "image", "Image", "主图链接"]:
                            image_col = col_idx
                
                if sku_col and image_col:
                    self._log_image(f"[i] Found columns in '{sheet_name}': SKU=Col{sku_col}, Image=Col{image_col}")
                    
                    # Read data rows
                    for row_idx in range(2, sheet.max_row + 1):
                        sku_cell = sheet.cell(row_idx, sku_col)
                        image_cell = sheet.cell(row_idx, image_col)
                        
                        if sku_cell.value and image_cell.value:
                            sku = str(sku_cell.value).strip()
                            image_url = str(image_cell.value).strip()
                            
                            # Validate URL
                            if image_url.startswith(("http://", "https://")):
                                results.append((sku, image_url))
            
            wb.close()
            
        except Exception as e:
            self._log_image(f"[FAIL] Error parsing {excel_file.name}: {str(e)}")
        
        return results
    
    def _browse_file(self):
        """Browse for Excel file"""
        filetypes = [
            ("Excel files", "*.xlsx *.xls"),
            ("All files", "*.*")
        ]
        
        filename = filedialog.askopenfilename(
            title="Select Excel file with accessory data",
            filetypes=filetypes
        )
        
        if filename:
            self.selected_file = Path(filename)
            self.file_path_var.set(str(self.selected_file))
            self._load_sheet_names()
            self.status_var.set(f"File selected: {self.selected_file.name}")
    
    def _load_sheet_names(self):
        """Load sheet names from Excel file"""
        if not self.selected_file or not self.selected_file.exists():
            return
        
        try:
            if OPENPYXL_AVAILABLE:
                wb = load_workbook(self.selected_file, read_only=True)
                sheet_names = wb.sheetnames
                wb.close()
            else:
                # Use zipfile method as fallback
                sheet_names = self._get_sheet_names_zipfile(self.selected_file)
            
            self.sheet_combo['values'] = sheet_names
            
            # Auto-select if "关联辅料" exists
            if "关联辅料" in sheet_names:
                self.sheet_var.set("关联辅料")
            elif sheet_names:
                self.sheet_var.set(sheet_names[0])
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load sheet names: {e}")
    
    def _get_sheet_names_zipfile(self, xlsx_path):
        """Get sheet names using zipfile (fallback method)"""
        try:
            with zipfile.ZipFile(xlsx_path) as z:
                workbook_xml = z.read('xl/workbook.xml')
                root = ET.fromstring(workbook_xml)
                sheets = root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet')
                return [sheet.get('name') for sheet in sheets]
        except:
            return ['Sheet1']  # Default fallback
    
    def _process_file(self):
        """Process the selected Excel file"""
        if not self.selected_file or not self.selected_file.exists():
            messagebox.showwarning("Warning", "Please select an Excel file first")
            return
        
        sheet_name = self.sheet_var.get()
        if not sheet_name:
            messagebox.showwarning("Warning", "Please specify a sheet name")
            return
        
        try:
            self.status_var.set("Processing Excel file...")
            self.root.update()
            
            # Read Excel data
            if OPENPYXL_AVAILABLE:
                self.excel_data = self._read_excel_openpyxl(self.selected_file, sheet_name)
            else:
                self.excel_data = self._read_excel_zipfile(self.selected_file, sheet_name)
            
            if not self.excel_data:
                messagebox.showwarning("Warning", f"No data found in sheet '{sheet_name}'")
                return
            
            # Update column combo boxes
            if self.excel_data:
                headers = self.excel_data[0]
                for i in range(5):  # Update all 5 column combos
                    combo = getattr(self, f"col_combo_{i}")
                    combo['values'] = headers
            
            # Display data preview
            self._update_data_preview()
            
            self.status_var.set(f"Processed {len(self.excel_data)} rows from {sheet_name}")
            
        except Exception as e:
            error_msg = f"Error processing file: {e}\n\n{traceback.format_exc()}"
            messagebox.showerror("Error", error_msg)
            self.status_var.set("Error processing file")
    
    def _read_excel_openpyxl(self, file_path, sheet_name):
        """Read Excel data using openpyxl"""
        wb = load_workbook(file_path, read_only=True, data_only=True)
        
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"Sheet '{sheet_name}' not found")
        
        ws = wb[sheet_name]
        data = []
        
        for row in ws.iter_rows(values_only=True):
            # Convert None values to empty strings
            row_data = [str(cell) if cell is not None else "" for cell in row]
            data.append(row_data)
        
        wb.close()
        return data
    
    def _read_excel_zipfile(self, file_path, sheet_name):
        """Read Excel data using zipfile method (fallback)"""
        # This is a simplified version - you may need to enhance for specific sheet selection
        with zipfile.ZipFile(file_path) as z:
            # Read shared strings
            shared = []
            if 'xl/sharedStrings.xml' in z.namelist():
                root = ET.fromstring(z.read('xl/sharedStrings.xml'))
                for si in root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}si'):
                    text = ''.join(t.text or '' for t in si.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t'))
                    shared.append(text)
            
            # Read first worksheet (simplified)
            sheet_root = ET.fromstring(z.read('xl/worksheets/sheet1.xml'))
            data = []
            
            for row in sheet_root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}row'):
                row_data = []
                for c in row.findall('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c'):
                    t = c.get('t')
                    if t == 's':
                        v = c.find('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v')
                        val = shared[int(v.text)] if v is not None and v.text else ''
                    else:
                        v = c.find('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v')
                        val = v.text if v is not None else ''
                    row_data.append(val)
                data.append(row_data)
            
            return data
    
    def _update_data_preview(self):
        """Update the data preview tree"""
        # Clear existing items
        for item in self.data_tree.get_children():
            self.data_tree.delete(item)
        
        if not self.excel_data or len(self.excel_data) < 2:
            return
        
        # Show first 20 rows (excluding header)
        for row in self.excel_data[1:21]:  # Skip header, show max 20 rows
            # Pad row to ensure it has enough columns
            padded_row = row + [''] * (5 - len(row))
            self.data_tree.insert('', 'end', values=padded_row[:5])
    
    def _generate_preview(self):
        """Generate preview of changes"""
        if not self.excel_data:
            messagebox.showwarning("Warning", "No Excel data loaded. Please process a file first.")
            return
        
        try:
            self.status_var.set("Generating preview...")
            self.root.update()
            
            # Get column mappings
            col_mapping = {}
            for label, var in self.column_vars.items():
                selected_col = var.get()
                if selected_col and self.excel_data:
                    try:
                        col_mapping[label] = self.excel_data[0].index(selected_col)
                    except ValueError:
                        messagebox.showerror("Error", f"Column '{selected_col}' not found in data")
                        return
            
            # Generate new mapping
            self.new_mapping = self._build_new_mapping(col_mapping)
            
            # Create preview text
            preview_lines = []
            preview_lines.append("ACCESSORY MAPPING UPDATE PREVIEW")
            preview_lines.append("=" * 50)
            preview_lines.append(f"Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            preview_lines.append(f"Source file: {self.selected_file.name if self.selected_file else 'Unknown'}")
            preview_lines.append(f"Data rows processed: {len(self.excel_data) - 1 if self.excel_data else 0}")
            preview_lines.append("")
            
            # Compare with current mapping
            current_products = set(self.current_mapping.get("products", {}).keys())
            new_products = set(self.new_mapping.get("products", {}).keys())
            
            added_products = new_products - current_products
            updated_products = current_products & new_products
            
            preview_lines.append(f"SUMMARY (ADD/UPDATE ONLY):")
            preview_lines.append(f"  Products to add: {len(added_products)}")
            preview_lines.append(f"  Products to update: {len(updated_products)}")
            preview_lines.append(f"  Existing products preserved: {len(current_products) - len(updated_products)}")
            preview_lines.append("")
            
            # Show details
            if added_products:
                preview_lines.append("NEW PRODUCTS TO ADD:")
                for sku in sorted(added_products)[:10]:  # Show first 10
                    product = self.new_mapping["products"][sku]
                    acc_count = len(product.get("accessories", []))
                    preview_lines.append(f"  + {sku}: {product.get('name', 'No name')} ({acc_count} accessories)")
                if len(added_products) > 10:
                    preview_lines.append(f"  ... and {len(added_products) - 10} more")
                preview_lines.append("")
            
            if updated_products:
                preview_lines.append("PRODUCTS TO UPDATE (first 10):")
                for sku in sorted(updated_products)[:10]:
                    product = self.new_mapping["products"][sku]
                    acc_count = len(product.get("accessories", []))
                    preview_lines.append(f"  * {sku}: {product.get('name', 'No name')} ({acc_count} accessories)")
                if len(updated_products) > 10:
                    preview_lines.append(f"  ... and {len(updated_products) - 10} more")
                preview_lines.append("")
            
            # Show sample of new mapping
            preview_lines.append("SAMPLE OF NEW MAPPING:")
            sample_count = 0
            for sku, product in self.new_mapping.get("products", {}).items():
                if sample_count >= 5:
                    break
                preview_lines.append(f"  {sku}:")
                preview_lines.append(f"    Name: {product.get('name', 'No name')}")
                for acc in product.get("accessories", [])[:3]:  # Show first 3 accessories
                    preview_lines.append(f"    Accessory: {acc.get('sku', 'No SKU')} ({acc.get('ratio_main', '1')}:{acc.get('ratio_accessory', '1')})")
                if len(product.get("accessories", [])) > 3:
                    preview_lines.append(f"    ... and {len(product.get('accessories', [])) - 3} more accessories")
                preview_lines.append("")
                sample_count += 1
            
            # Display preview
            self.preview_text.delete(1.0, tk.END)
            self.preview_text.insert(1.0, "\n".join(preview_lines))
            
            self.status_var.set("Preview generated successfully")
            
        except Exception as e:
            error_msg = f"Error generating preview: {e}\n\n{traceback.format_exc()}"
            messagebox.showerror("Error", error_msg)
            self.status_var.set("Error generating preview")
    
    def _build_new_mapping(self, col_mapping):
        """Build new mapping from Excel data"""
        if not self.excel_data or len(self.excel_data) < 2:
            return {"products": {}}
        
        # Get required column indices
        required_cols = ["主产品SKU", "辅料SKU"]
        for col in required_cols:
            if col not in col_mapping:
                raise ValueError(f"Required column '{col}' not mapped")
        
        main_sku_idx = col_mapping["主产品SKU"]
        acc_sku_idx = col_mapping["辅料SKU"]
        acc_name_idx = col_mapping.get("辅料名称")
        main_qty_idx = col_mapping.get("主产品数量")
        acc_qty_idx = col_mapping.get("辅料数量")
        
        # Build mapping
        products = {}
        
        for row in self.excel_data[1:]:  # Skip header
            if len(row) <= max(main_sku_idx, acc_sku_idx):
                continue
            
            main_sku = str(row[main_sku_idx]).strip()
            acc_sku = str(row[acc_sku_idx]).strip()
            
            if not main_sku or not acc_sku:
                continue
            
            # Get accessory details
            acc_name = str(row[acc_name_idx]).strip() if acc_name_idx and len(row) > acc_name_idx else ""
            main_qty = str(row[main_qty_idx]).strip() if main_qty_idx and len(row) > main_qty_idx else "1"
            acc_qty = str(row[acc_qty_idx]).strip() if acc_qty_idx and len(row) > acc_qty_idx else "1"
            
            # Ensure main product exists
            if main_sku not in products:
                # Try to get product name from current mapping
                current_product = self.current_mapping.get("products", {}).get(main_sku, {})
                products[main_sku] = {
                    "name": current_product.get("name", ""),
                    "accessories": []
                }
            
            # Add accessory
            accessory = {
                "sku": acc_sku,
                "name": acc_name,
                "ratio_main": main_qty,
                "ratio_accessory": acc_qty
            }
            
            # Check if this accessory already exists for this product
            existing_accessories = products[main_sku]["accessories"]
            if not any(acc["sku"] == acc_sku for acc in existing_accessories):
                existing_accessories.append(accessory)
        
        return {"products": products}
    
    def _create_backup(self):
        """Create backup of current mapping"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = self.mapping_file.parent / f"accessory_mapping_backup_{timestamp}.json"
            
            with open(backup_path, 'w', encoding='utf-8') as f:
                json.dump(self.current_mapping, f, ensure_ascii=False, indent=2)
            
            messagebox.showinfo("Success", f"Backup created: {backup_path.name}")
            self.status_var.set(f"Backup created: {backup_path.name}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to create backup: {e}")
    
    def _apply_changes(self):
        """Apply the new mapping - merges with existing (add/update only, no removal)"""
        if not self.new_mapping:
            messagebox.showwarning("Warning", "No new mapping generated. Please generate preview first.")
            return
        
        # Confirm with user
        result = messagebox.askyesno(
            "Confirm Changes",
            "This will ADD/UPDATE products in the accessory mapping file.\n"
            "Existing products not in this file will be preserved.\n\n"
            "Do you want to continue?"
        )
        
        if not result:
            return
        
        try:
            # Merge new mapping with current mapping (add/update only)
            merged_mapping = self.current_mapping.copy()
            if "products" not in merged_mapping:
                merged_mapping["products"] = {}
            
            new_products = self.new_mapping.get("products", {})
            
            for product_sku, product_data in new_products.items():
                if product_sku in merged_mapping["products"]:
                    # Update existing product - merge accessories
                    existing_product = merged_mapping["products"][product_sku]
                    
                    # Update name if provided and not empty
                    if product_data.get("name"):
                        existing_product["name"] = product_data["name"]
                    
                    # Merge accessories - update existing, add new
                    existing_accessories = {acc["sku"]: acc for acc in existing_product.get("accessories", [])}
                    new_accessories = product_data.get("accessories", [])
                    
                    for new_acc in new_accessories:
                        acc_sku = new_acc["sku"]
                        existing_accessories[acc_sku] = new_acc
                    
                    existing_product["accessories"] = list(existing_accessories.values())
                else:
                    # Add new product
                    merged_mapping["products"][product_sku] = product_data
            
            # Write merged mapping
            with open(self.mapping_file, 'w', encoding='utf-8') as f:
                json.dump(merged_mapping, f, ensure_ascii=False, indent=2)
            
            # Update current mapping
            self.current_mapping = merged_mapping
            
            # Refresh displays
            self._refresh_current_mapping()
            
            updated_count = len(new_products)
            messagebox.showinfo("Success", 
                              f"Accessory mapping updated successfully!\n\n"
                              f"Updated/Added {updated_count} product(s)")
            self.status_var.set(f"Mapping updated: {updated_count} products")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to apply changes: {e}")
    
    def _refresh_current_mapping(self):
        """Refresh the current mapping display"""
        # Clear existing items
        for item in self.mapping_tree.get_children():
            self.mapping_tree.delete(item)
        
        # Reload mapping
        self._load_current_mapping()
        
        # Populate tree
        for product_sku, product_info in self.current_mapping.get("products", {}).items():
            product_name = product_info.get("name", "")
            accessories = product_info.get("accessories", [])
            
            if accessories:
                for acc in accessories:
                    self.mapping_tree.insert('', 'end', values=(
                        product_sku,
                        product_name,
                        acc.get("sku", ""),
                        acc.get("name", ""),
                        acc.get("ratio_main", "1"),
                        acc.get("ratio_accessory", "1")
                    ))
            else:
                # Show product even if no accessories
                self.mapping_tree.insert('', 'end', values=(
                    product_sku,
                    product_name,
                    "",
                    "",
                    "",
                    ""
                ))
        
        # Update status
        product_count = len(self.current_mapping.get("products", {}))
        total_accessories = sum(len(p.get("accessories", [])) for p in self.current_mapping.get("products", {}).values())
        self.status_var.set(f"Current mapping: {product_count} products, {total_accessories} accessory relationships")
    
    def _filter_current_mapping(self, *args):
        """Filter the current mapping display based on search"""
        search_term = self.search_var.get().lower()
        
        # Clear and repopulate with filtered results
        for item in self.mapping_tree.get_children():
            self.mapping_tree.delete(item)
        
        for product_sku, product_info in self.current_mapping.get("products", {}).items():
            product_name = product_info.get("name", "").lower()
            
            # Check if search term matches product SKU or name
            if search_term in product_sku.lower() or search_term in product_name:
                accessories = product_info.get("accessories", [])
                
                if accessories:
                    for acc in accessories:
                        self.mapping_tree.insert('', 'end', values=(
                            product_sku,
                            product_info.get("name", ""),
                            acc.get("sku", ""),
                            acc.get("name", ""),
                            acc.get("ratio_main", "1"),
                            acc.get("ratio_accessory", "1")
                        ))
                else:
                    self.mapping_tree.insert('', 'end', values=(
                        product_sku,
                        product_info.get("name", ""),
                        "",
                        "",
                        "",
                        ""
                    ))


def main():
    """Main function to run the GUI"""
    if not OPENPYXL_AVAILABLE:
        print("Warning: openpyxl not available. Some features may be limited.")
        print("Install it with: pip install openpyxl")
    
    root = tk.Tk()
    app = AccessoryMappingUpdaterGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
