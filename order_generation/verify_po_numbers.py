#!/usr/bin/env python3
"""Verify that PO import files have dashes instead of underscores in 采购单号."""

import openpyxl
from pathlib import Path

def verify_po_numbers():
    """Check a sample PO import file for underscore replacement."""
    
    po_file = Path(__file__).parent / "PO_import_filled" / "PO_import_verify_Elasticbrush01.xlsx"
    
    if not po_file.exists():
        print(f"Error: File not found: {po_file}")
        return False
    
    wb = openpyxl.load_workbook(po_file)
    ws = wb.active
    
    # Find 采购单号 column
    headers = []
    for col in range(1, ws.max_column + 1):
        header = ws.cell(2, col).value
        headers.append(header)
    
    if '采购单号' not in headers:
        print("Error: 采购单号 column not found")
        return False
    
    po_col = headers.index('采购单号') + 1
    
    print("Checking 采购单号 values in PO import file:")
    print("=" * 70)
    
    has_underscore = False
    sample_values = []
    
    for row in range(3, min(10, ws.max_row + 1)):
        value = ws.cell(row, po_col).value
        if value:
            sample_values.append(value)
            if '_' in str(value):
                has_underscore = True
                print(f"  Row {row}: {value} [ERROR: Contains underscore]")
            else:
                print(f"  Row {row}: {value} [OK]")
    
    print("=" * 70)
    
    if has_underscore:
        print("\n[ERROR] Some 采购单号 values still contain underscores!")
        return False
    else:
        print("\n[SUCCESS] All 采购单号 values use dashes instead of underscores!")
        print(f"Sample values: {set(sample_values)}")
        return True

if __name__ == "__main__":
    success = verify_po_numbers()
    exit(0 if success else 1)
