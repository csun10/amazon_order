#!/usr/bin/env python3
"""Verification script to test order generation for all SKUs with sales > 10 in 30 days.

This script:
1. Reads the listing file to find SKUs with 30-day sales > 10
2. Generates orders for each SKU one by one (test quantity of 100)
3. Saves Excel and PO_import outputs for validation
4. Logs all results
"""

import openpyxl
from pathlib import Path
import json
import sys
import traceback
from datetime import datetime

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent))

from direct_sku_to_json import generate_factory_jsons
from fill_po_import import fill_po_import_for_order

# Paths
ROOT = Path(__file__).resolve().parent
LISTING_FILE = ROOT / "docs" / "Listing20260203-877196087228878848.xlsx"
VERIFICATION_OUTPUT = ROOT / "verification_output"
VERIFICATION_LOG = VERIFICATION_OUTPUT / "verification_log.txt"

# Create verification output directory
VERIFICATION_OUTPUT.mkdir(exist_ok=True)

def log_message(message, log_file=VERIFICATION_LOG):
    """Log message to both console and file."""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_line = f"[{timestamp}] {message}"
    print(log_line)
    with open(log_file, 'a', encoding='utf-8') as f:
        f.write(log_line + '\n')

def read_listing_file(xlsx_path: Path, min_sales: int = 10) -> list:
    """Read listing file and extract SKUs with 30-day sales > min_sales.
    
    Returns:
        List of tuples: (SKU, MSKU, 产品名称, 30天销量)
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    
    skus = []
    
    # Headers are in row 1
    # Column 1: MSKU
    # Column 7: 产品名称 (品名)
    # Column 10: SKU
    # Column 13: 30天销量
    
    for row in range(2, ws.max_row + 1):
        msku = ws.cell(row, 1).value
        product_name = ws.cell(row, 9).value  # Column 9: 品名
        sku = ws.cell(row, 10).value
        sales_30d = ws.cell(row, 13).value
        
        # Skip if no SKU or sales data
        if not sku or sales_30d is None:
            continue
        
        try:
            sales_30d = int(sales_30d)
        except (ValueError, TypeError):
            continue
        
        if sales_30d > min_sales:
            skus.append({
                'sku': str(sku).strip(),
                'msku': str(msku).strip() if msku else '',
                'product_name': str(product_name).strip() if product_name else '',
                'sales_30d': sales_30d
            })
    
    wb.close()
    return skus

def check_json_template_exists(sku: str) -> bool:
    """Check if JSON template exists for SKU."""
    template_path = ROOT / "json_template" / f"{sku}.json"
    return template_path.exists()

def verify_single_sku(sku: str, product_info: dict, test_qty: int = 100) -> dict:
    """
    Verify order generation for a single SKU.
    
    Returns:
        dict with results: {
            'sku': str,
            'success': bool,
            'has_template': bool,
            'excel_files': list,
            'po_import_file': str or None,
            'error': str or None
        }
    """
    result = {
        'sku': sku,
        'msku': product_info['msku'],
        'product_name': product_info['product_name'],
        'sales_30d': product_info['sales_30d'],
        'success': False,
        'has_template': False,
        'excel_files': [],
        'po_import_file': None,
        'error': None
    }
    
    # Check if template exists
    if not check_json_template_exists(sku):
        result['error'] = 'No JSON template found'
        return result
    
    result['has_template'] = True
    
    try:
        order_name = f"verify_{sku}"
        
        # Step 1: Generate factory-grouped JSON and Excel files
        log_message(f"  Processing {sku} (Qty: {test_qty})...")
        
        # Create SKU/quantity pairs dictionary
        pairs = {sku: test_qty}
        
        # Generate factory JSONs (this also creates Excel files automatically)
        json_files = generate_factory_jsons(pairs, order_name)
        
        if not json_files:
            result['error'] = 'No JSON files generated'
            return result
        
        log_message(f"    Generated {len(json_files)} factory group(s)")
        
        # Step 2: Copy outputs to verification directory
        excel_output_dir = VERIFICATION_OUTPUT / "excel"
        excel_output_dir.mkdir(exist_ok=True)
        
        import shutil
        
        # Copy Excel files
        excel_export_dir = ROOT / "PO_excel_export"
        for json_file in json_files:
            excel_filename = json_file.stem + '.xlsx'
            excel_path = excel_export_dir / excel_filename
            
            if excel_path.exists():
                new_excel_path = excel_output_dir / excel_filename
                shutil.copy2(excel_path, new_excel_path)
                result['excel_files'].append(str(new_excel_path))
                log_message(f"    Copied Excel: {excel_filename}")
        
        # Step 3: Generate PO import
        po_import_dir = VERIFICATION_OUTPUT / "po_import"
        po_import_dir.mkdir(exist_ok=True)
        
        try:
            po_import_path = fill_po_import_for_order(order_name)
            
            # Move to verification directory
            new_po_path = po_import_dir / po_import_path.name
            if po_import_path.exists():
                shutil.copy2(po_import_path, new_po_path)
                result['po_import_file'] = str(new_po_path)
                log_message(f"    Created PO Import: {new_po_path.name}")
        except Exception as e:
            log_message(f"    Warning: PO import generation failed: {e}")
        
        result['success'] = True
        
    except Exception as e:
        result['error'] = str(e)
        log_message(f"    ERROR: {e}")
        traceback.print_exc()
    
    return result

def main():
    """Main verification function."""
    log_message("="*80)
    log_message("STARTING ORDER GENERATION VERIFICATION")
    log_message("="*80)
    
    # Check listing file exists
    if not LISTING_FILE.exists():
        log_message(f"ERROR: Listing file not found: {LISTING_FILE}")
        return 1
    
    log_message(f"Reading listing file: {LISTING_FILE.name}")
    
    # Read SKUs with sales > 10
    skus = read_listing_file(LISTING_FILE, min_sales=10)
    
    log_message(f"Found {len(skus)} SKUs with 30-day sales > 10")
    
    if not skus:
        log_message("No SKUs to process.")
        return 0
    
    # Statistics
    results = []
    success_count = 0
    failed_count = 0
    no_template_count = 0
    
    # Process each SKU
    log_message("\n" + "="*80)
    log_message("PROCESSING SKUs")
    log_message("="*80)
    
    for idx, sku_info in enumerate(skus, start=1):
        sku = sku_info['sku']
        log_message(f"\n[{idx}/{len(skus)}] Processing SKU: {sku}")
        log_message(f"  MSKU: {sku_info['msku']}")
        log_message(f"  Product: {sku_info['product_name']}")
        log_message(f"  30-day Sales: {sku_info['sales_30d']}")
        
        result = verify_single_sku(sku, sku_info, test_qty=100)
        results.append(result)
        
        if result['success']:
            success_count += 1
            log_message(f"  ✓ SUCCESS")
        else:
            failed_count += 1
            if not result['has_template']:
                no_template_count += 1
            log_message(f"  ✗ FAILED: {result['error']}")
    
    # Generate summary report
    log_message("\n" + "="*80)
    log_message("VERIFICATION SUMMARY")
    log_message("="*80)
    log_message(f"Total SKUs processed: {len(skus)}")
    log_message(f"  ✓ Successful: {success_count}")
    log_message(f"  ✗ Failed: {failed_count}")
    log_message(f"    - No template: {no_template_count}")
    log_message(f"    - Other errors: {failed_count - no_template_count}")
    
    # Save detailed results to JSON
    results_file = VERIFICATION_OUTPUT / "verification_results.json"
    with open(results_file, 'w', encoding='utf-8') as f:
        json.dump({
            'timestamp': datetime.now().isoformat(),
            'listing_file': str(LISTING_FILE),
            'total_skus': len(skus),
            'success_count': success_count,
            'failed_count': failed_count,
            'no_template_count': no_template_count,
            'results': results
        }, f, indent=2, ensure_ascii=False)
    
    log_message(f"\nDetailed results saved to: {results_file}")
    
    # Print failed SKUs
    if failed_count > 0:
        log_message("\n" + "="*80)
        log_message("FAILED SKUs")
        log_message("="*80)
        
        for result in results:
            if not result['success']:
                log_message(f"  - {result['sku']}: {result['error']}")
    
    log_message("\n" + "="*80)
    log_message("VERIFICATION COMPLETE")
    log_message("="*80)
    log_message(f"Output directory: {VERIFICATION_OUTPUT}")
    log_message(f"  - Excel files: {VERIFICATION_OUTPUT / 'excel'}")
    log_message(f"  - PO imports: {VERIFICATION_OUTPUT / 'po_import'}")
    log_message(f"  - Log file: {VERIFICATION_LOG}")
    log_message(f"  - Results: {results_file}")
    
    return 0 if failed_count == 0 else 1

if __name__ == "__main__":
    sys.exit(main())
