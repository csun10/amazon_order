# -*- coding: utf-8 -*-
"""
Quick Pipeline Test

Tests the entire order generation pipeline:
1. Excel Template → JSON
2. JSON → Excel Output
3. PO Import Generation
4. Data Consistency Verification
"""
import sys
import io
from pathlib import Path
import json
import subprocess
from openpyxl import load_workbook
from datetime import datetime

# Force UTF-8 output
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

ROOT = Path(__file__).resolve().parent
TEST_SKU = "ST1122-1"  # Use a parent product for testing

class Colors:
    GREEN = '\033[92m'
    RED = '\033[91m'
    YELLOW = '\033[93m'
    BLUE = '\033[94m'
    RESET = '\033[0m'
    BOLD = '\033[1m'

def print_section(title):
    print(f"\n{'='*70}")
    print(f"{Colors.BOLD}{title}{Colors.RESET}")
    print(f"{'='*70}\n")

def print_ok(msg):
    print(f"{Colors.GREEN}[OK]{Colors.RESET} {msg}")

def print_fail(msg):
    print(f"{Colors.RED}[FAIL]{Colors.RESET} {msg}")

def print_info(msg):
    print(f"{Colors.BLUE}[INFO]{Colors.RESET} {msg}")

def print_warn(msg):
    print(f"{Colors.YELLOW}[WARN]{Colors.RESET} {msg}")

def test_excel_to_json():
    """Test 1: Excel Template → JSON"""
    print_section("TEST 1: Excel Template → JSON Conversion")
    
    excel_path = ROOT / "PO_excel_template" / f"{TEST_SKU}.xlsx"
    json_path = ROOT / "json_template" / f"{TEST_SKU}.json"
    
    if not excel_path.exists():
        print_fail(f"Excel template not found: {excel_path}")
        return None
    
    print_info(f"Source: {excel_path.name}")
    print_info(f"Target: {json_path.name}")
    
    try:
        # Read Excel
        wb = load_workbook(excel_path)
        ws = wb.active
        
        excel_data = {
            'supplier': ws['B3'].value,
            'buyer': ws['B69'].value,
            'first_product': ws['A7'].value,
        }
        
        print_info(f"Excel B3 (Supplier): {excel_data['supplier']}")
        print_info(f"Excel B69 (Buyer): {excel_data['buyer']}")
        print_info(f"Excel A7 (First SKU): {excel_data['first_product']}")
        
        wb.close()
        
        # Check JSON exists and is recent
        if not json_path.exists():
            print_warn("JSON template doesn't exist - run excel_to_json_template.py")
            return None
        
        # Read JSON
        with open(json_path, 'r', encoding='utf-8-sig') as f:
            json_data = json.load(f)
        
        json_supplier = json_data.get('cells', {}).get('B3', {}).get('value')
        json_buyer = json_data.get('footer', {}).get('buyer')
        json_products = json_data.get('products', [])
        json_first_sku = json_products[0].get('产品编号') if json_products else None
        
        print_info(f"JSON B3 (Supplier): {json_supplier}")
        print_info(f"JSON B69 (Buyer): {json_buyer}")
        print_info(f"JSON First Product: {json_first_sku}")
        
        # Verify consistency
        issues = []
        if excel_data['supplier'] != json_supplier:
            issues.append(f"Supplier mismatch: Excel={excel_data['supplier']} vs JSON={json_supplier}")
        
        if excel_data['buyer'] != json_buyer:
            issues.append(f"Buyer mismatch: Excel={excel_data['buyer']} vs JSON={json_buyer}")
        
        if excel_data['first_product'] != json_first_sku:
            issues.append(f"Product SKU mismatch: Excel={excel_data['first_product']} vs JSON={json_first_sku}")
        
        if issues:
            for issue in issues:
                print_fail(issue)
            print_warn("Run: python excel_to_json_template.py to sync")
            return None
        else:
            print_ok("Excel → JSON data is consistent!")
            return {
                'excel': excel_data,
                'json': json_data,
                'json_path': json_path
            }
    
    except Exception as e:
        print_fail(f"Error: {e}")
        return None

def test_json_to_excel(json_data_result):
    """Test 2: JSON → Excel Output"""
    print_section("TEST 2: JSON → Excel Output Generation")
    
    if not json_data_result:
        print_fail("Skipped - previous test failed")
        return None
    
    json_path = json_data_result['json_path']
    test_name = f"test_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    output_path = ROOT / "PO_excel_export" / f"{test_name}.xlsx"
    
    # Copy JSON to json_exports for PO import processing
    # PO import expects pattern: {name}-{number}.json
    json_export_path = ROOT / "json_exports" / f"{test_name}-1.json"
    json_export_path.parent.mkdir(exist_ok=True)
    
    import shutil
    shutil.copy(json_path, json_export_path)
    print_info(f"Copied JSON to: {json_export_path.name}")
    
    print_info(f"Input: {json_path.name}")
    print_info(f"Output: {output_path.name}")
    
    try:
        # Run json_PO_excel.py
        result = subprocess.run([
            sys.executable,
            str(ROOT / "json_PO_excel.py"),
            str(json_export_path),
            str(output_path)
        ], capture_output=True, text=True, encoding='utf-8')
        
        if result.returncode != 0:
            print_fail(f"json_PO_excel.py failed: {result.stderr}")
            return None
        
        if not output_path.exists():
            print_fail(f"Output file not created: {output_path}")
            return None
        
        # Read generated Excel
        wb = load_workbook(output_path)
        ws = wb.active
        
        output_data = {
            'supplier': ws['E69'].value,
            'buyer': ws['B69'].value,
            'first_product': ws['A7'].value,
        }
        
        print_info(f"Output E69 (Supplier): {output_data['supplier']}")
        print_info(f"Output B69 (Buyer): {output_data['buyer']}")
        print_info(f"Output A7 (First SKU): {output_data['first_product']}")
        
        wb.close()
        
        # Verify consistency with JSON
        json_supplier = json_data_result['json'].get('cells', {}).get('B3', {}).get('value')
        json_buyer = json_data_result['json'].get('footer', {}).get('buyer')
        
        issues = []
        if output_data['supplier'] != json_supplier:
            issues.append(f"Supplier mismatch: Output={output_data['supplier']} vs JSON={json_supplier}")
        
        if output_data['buyer'] != json_buyer:
            issues.append(f"Buyer mismatch: Output={output_data['buyer']} vs JSON={json_buyer}")
        
        if issues:
            for issue in issues:
                print_fail(issue)
            return None
        else:
            print_ok("JSON → Excel Output data is consistent!")
            return {
                'output': output_data,
                'output_path': output_path,
                'test_name': test_name,
                'json_export_path': json_export_path
            }
    
    except Exception as e:
        print_fail(f"Error: {e}")
        return None

def test_po_import(json_data_result, excel_output_result):
    """Test 3: PO Import Generation"""
    print_section("TEST 3: PO Import Generation")
    
    if not json_data_result or not excel_output_result:
        print_fail("Skipped - previous tests failed")
        return None
    
    test_name = excel_output_result['test_name']
    po_import_path = ROOT / "PO_import_filled" / f"PO_import_{test_name}.xlsx"
    
    print_info(f"Test order: {test_name}")
    print_info(f"Expected output: {po_import_path.name}")
    
    try:
        # Run fill_po_import.py
        result = subprocess.run([
            sys.executable,
            str(ROOT / "fill_po_import.py"),
            test_name
        ], capture_output=True, text=True, encoding='utf-8')
        
        if result.returncode != 0:
            print_fail(f"fill_po_import.py failed: {result.stderr}")
            return None
        
        if not po_import_path.exists():
            print_fail(f"PO import file not created: {po_import_path}")
            return None
        
        # Read PO import
        wb = load_workbook(po_import_path)
        ws = wb.active
        
        # Find first data row (row 3)
        po_data = {
            'supplier': ws.cell(3, 3).value,  # Column 3 = *供应商
            'buyer': ws.cell(3, 5).value,     # Column 5 = 采购方
            'sku': ws.cell(3, 25).value,      # Column 25 = *SKU
        }
        
        print_info(f"PO Import Row 3:")
        print_info(f"  供应商: {po_data['supplier']}")
        print_info(f"  采购方: {po_data['buyer']}")
        print_info(f"  SKU: {po_data['sku']}")
        
        wb.close()
        
        # Verify consistency
        json_supplier = json_data_result['json'].get('cells', {}).get('B3', {}).get('value')
        json_buyer = json_data_result['json'].get('footer', {}).get('buyer')
        
        issues = []
        if po_data['supplier'] != json_supplier:
            issues.append(f"Supplier mismatch: PO={po_data['supplier']} vs JSON={json_supplier}")
        
        # Only check buyer for parent products
        if po_data['buyer'] and po_data['buyer'] != json_buyer:
            issues.append(f"Buyer mismatch: PO={po_data['buyer']} vs JSON={json_buyer}")
        
        if po_data['sku'] != TEST_SKU:
            issues.append(f"SKU mismatch: PO={po_data['sku']} vs Expected={TEST_SKU}")
        
        if issues:
            for issue in issues:
                print_fail(issue)
            return None
        else:
            print_ok("PO Import data is consistent!")
            return {
                'po_import': po_data,
                'po_import_path': po_import_path
            }
    
    except Exception as e:
        print_fail(f"Error: {e}")
        import traceback
        traceback.print_exc()
        return None

def cleanup_test_files(excel_output_result, po_import_result):
    """Clean up test files"""
    print_section("CLEANUP")
    
    files_to_delete = []
    
    if excel_output_result:
        # Delete test JSON export
        if 'json_export_path' in excel_output_result and excel_output_result['json_export_path'].exists():
            files_to_delete.append(excel_output_result['json_export_path'])
        
        # Delete test Excel output
        if excel_output_result['output_path'].exists():
            files_to_delete.append(excel_output_result['output_path'])
    
    if po_import_result and po_import_result['po_import_path'].exists():
        files_to_delete.append(po_import_result['po_import_path'])
    
    for file in files_to_delete:
        try:
            file.unlink()
            print_info(f"Deleted: {file.name}")
        except Exception as e:
            print_warn(f"Could not delete {file.name}: {e}")
    
    print_ok(f"Cleaned up {len(files_to_delete)} test files")

def main():
    print(f"\n{Colors.BOLD}Testing with SKU: {TEST_SKU}{Colors.RESET}")
    
    # Test 1: Excel → JSON
    json_result = test_excel_to_json()
    
    # Test 2: JSON → Excel
    excel_result = test_json_to_excel(json_result)
    
    # Test 3: PO Import
    po_result = test_po_import(json_result, excel_result)
    
    # Summary
    print_section("SUMMARY")
    
    tests = [
        ("Excel → JSON", json_result),
        ("JSON → Excel Output", excel_result),
        ("PO Import Generation", po_result)
    ]
    
    passed = sum(1 for _, result in tests if result is not None)
    total = len(tests)
    
    for test_name, result in tests:
        if result:
            print_ok(f"{test_name}")
        else:
            print_fail(f"{test_name}")
    
    print(f"\n{Colors.BOLD}Result: {passed}/{total} tests passed{Colors.RESET}")
    
    if passed == total:
        print(f"\n{Colors.GREEN}{Colors.BOLD}✓ ALL TESTS PASSED - Pipeline is working correctly!{Colors.RESET}\n")
        
        # Cleanup
        cleanup_test_files(excel_result, po_result)
        
        return 0
    else:
        print(f"\n{Colors.RED}{Colors.BOLD}✗ SOME TESTS FAILED - Please review errors above{Colors.RESET}\n")
        return 1

if __name__ == "__main__":
    sys.exit(main())
