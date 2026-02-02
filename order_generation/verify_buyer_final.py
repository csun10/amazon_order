# -*- coding: utf-8 -*-
import sys, io
from openpyxl import load_workbook
from pathlib import Path

if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

ROOT = Path(__file__).resolve().parent

print("Checking Test2 order files...\n")

# Check Excel Output
excel_file = ROOT / "PO_excel_export" / "test2-1.xlsx"
if excel_file.exists():
    wb = load_workbook(excel_file)
    ws = wb.active
    print(f"[Excel Output] {excel_file.name}")
    print(f"  B69 (Buyer): {ws['B69'].value}")
    print(f"  E69 (Supplier): {ws['E69'].value}")
    print(f"  A7 (SKU): {ws['A7'].value}")
    wb.close()
else:
    print("[Excel Output] File not found")

print()

# Check PO Import
po_file = ROOT / "PO_import_filled" / "PO_import_test2.xlsx"
if po_file.exists():
    wb = load_workbook(po_file)
    ws = wb.active
    print(f"[PO Import] {po_file.name}")
    print(f"  Row 3 SKU (Col 25): {ws.cell(3, 25).value}")
    print(f"  Row 3 Buyer (Col 5): {ws.cell(3, 5).value or '(blank)'}")
    print(f"  Row 3 Supplier (Col 3): {ws.cell(3, 3).value}")
    wb.close()
else:
    print("[PO Import] File not found")

print("\n" + "="*60)
print("Expected for Elasticbrush01:")
print("  Buyer: 宁波集秀美容科技有限公司 (JIXIU)")
print("="*60)
