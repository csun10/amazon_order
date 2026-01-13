import openpyxl
import json
import os
from pathlib import Path

# First, identify all parent SKUs with 细针洗头刷
json_template_dir = Path(r"c:\Users\Cheng\Desktop\amazon_order\order_generation\json_template")
po_excel_dir = Path(r"c:\Users\Cheng\Desktop\amazon_order\order_generation\PO_excel")

# Search for parent SKUs containing 细针洗头刷
parent_skus = []

print("Searching for parent SKUs with 细针洗头刷...")
for json_file in json_template_dir.glob("*.json"):
    try:
        with open(json_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
            found = False
            
            # Search in cells
            for cell_key, cell_data in data.get('cells', {}).items():
                if isinstance(cell_data, dict):
                    text = cell_data.get('text', '') or cell_data.get('value', '')
                    if '细针洗头刷' in str(text):
                        found = True
                        break
            
            # Search in products array
            if not found:
                for product in data.get('products', []):
                    desc = product.get('描述', '')
                    if isinstance(desc, dict) and desc.get('type') == 'rich_text':
                        # Handle rich text content
                        for content in desc.get('content', []):
                            if '细针洗头刷' in str(content.get('text', '')):
                                found = True
                                break
                    elif '细针洗头刷' in str(desc):
                        found = True
                        break
                    if found:
                        break
            
            if found:
                parent_sku = json_file.stem
                if parent_sku not in parent_skus:
                    parent_skus.append(parent_sku)
                    print(f"Found: {parent_sku}")
    except Exception as e:
        print(f"Error reading {json_file.name}: {e}")

print(f"\nTotal parent SKUs found: {len(parent_skus)}")
print(f"Parent SKUs: {', '.join(parent_skus)}")

# Now update the PO Excel files
old_text_patterns = [
    "3）硬度按照55",
    "硬度55"
]
new_text_map = {
    "3）硬度按照55": "3）硬度按照55(伟氏硬度计D型测值65+-0.8)",
    "硬度55": "硬度55(伟氏硬度计D型测值65+-0.8)"
}

updated_files = []
not_found_files = []

print(f"\n{'='*60}")
print("Updating PO Excel files...")
print(f"{'='*60}")

for parent_sku in parent_skus:
    excel_file = po_excel_dir / f"{parent_sku}.xlsx"
    
    if not excel_file.exists():
        print(f"⚠️  Excel file not found: {excel_file.name}")
        not_found_files.append(parent_sku)
        continue
    
    try:
        wb = openpyxl.load_workbook(excel_file)
        found_and_replaced = False
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        for old_text in old_text_patterns:
                            if old_text in cell.value:
                                # Replace the text
                                new_text = new_text_map[old_text]
                                cell.value = cell.value.replace(old_text, new_text)
                                found_and_replaced = True
                                print(f"✓ Updated {excel_file.name} - Sheet: {sheet_name}, Cell: {cell.coordinate}")
        
        if found_and_replaced:
            wb.save(excel_file)
            updated_files.append(parent_sku)
            print(f"  Saved: {excel_file.name}")
        else:
            print(f"⚠️  Text not found in {excel_file.name}")
    
    except Exception as e:
        print(f"❌ Error processing {excel_file.name}: {e}")

print(f"\n{'='*60}")
print("Summary:")
print(f"{'='*60}")
print(f"Total parent SKUs identified: {len(parent_skus)}")
print(f"Successfully updated: {len(updated_files)}")
print(f"Excel files not found: {len(not_found_files)}")

if updated_files:
    print(f"\nUpdated files:")
    for sku in updated_files:
        print(f"  - {sku}.xlsx")

if not_found_files:
    print(f"\nExcel files not found:")
    for sku in not_found_files:
        print(f"  - {sku}.xlsx")
